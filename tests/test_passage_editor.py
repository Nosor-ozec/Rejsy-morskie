from __future__ import annotations

import hashlib
import json
import shutil
import threading
from pathlib import Path
from urllib.request import urlopen
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

import rejsy_morskie.passage_editor as passage_editor
from rejsy_morskie.passage_editor import (
    EDITOR_PROTOCOL_VERSION,
    _verification_differences,
    create_editor_server,
    load_passage_document,
    save_passage,
)


ROOT = Path(__file__).resolve().parents[1]
HEADERS = [
    "Nazwa", "Kraj", "Lat", "Lon", "Typ", "Uwagi", "Przejscie",
    "Przejscie_lp", "Przejscie_status",
]


def workdir() -> Path:
    path = Path(__file__).resolve().parent / "_work" / uuid4().hex
    path.mkdir(parents=True)
    return path


def make_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    locations = workbook.active
    locations.title = "Lokalizacje"
    locations.append(HEADERS)
    for row in rows:
        locations.append(row)
    other = workbook.create_sheet("Pozostale")
    other["A1"] = "nie zmieniaj"
    workbook.save(path)


def standard_rows() -> list[list[object]]:
    return [
        ["Port zwykły", "Polska", 54.0, 18.0, "Port", "zostaje", "", ""],
        ["Alfa 02", "Test", 2.0, 20.0, "Przejscie", "drugie", "Alfa", 2],
        ["Lokalizacja pośrodku", "Polska", 52.0, 19.0, "Wyspa", "też zostaje", "", ""],
        ["Alfa 01", "Test", 1.0, 10.0, "Przejscie", "pierwsze", "Alfa", 1],
        ["Beta 01", "", 3.0, 30.0, "", "", "Beta", 1],
        ["Beta 02", "", 4.0, 40.0, "", "", "Beta", 2],
    ]


def passage(document: dict[str, object], name: str) -> dict[str, object]:
    return next(item for item in document["passages"] if item["name"] == name)


def save(path: Path, name: str, points: list[dict[str, object]], original="Alfa"):
    document = load_passage_document(path)
    return save_passage(path, name, points, document["revision"], original)


def location_values(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        return list(workbook["Lokalizacje"].iter_rows(values_only=True))
    finally:
        workbook.close()


def test_reads_existing_passage_sorted_by_przejscie_lp():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    alfa = passage(load_passage_document(path), "Alfa")
    assert [point["name"] for point in alfa["points"]] == ["Alfa 01", "Alfa 02"]
    assert [point["lat"] for point in alfa["points"]] == [1.0, 2.0]


def test_adds_point_and_numbers_all_points_without_gaps():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    points = passage(load_passage_document(path), "Alfa")["points"]
    points.append({"lat": 2.5, "lon": 25.0, "locationType": "Przejscie"})
    result = save(path, "Alfa", points)
    alfa = passage(result, "Alfa")
    assert [point["name"] for point in alfa["points"]] == ["Alfa 01", "Alfa 02", "Alfa 03"]
    assert [point["order"] for point in alfa["points"]] == [1, 2, 3]
    assert path.with_name("rejsy.editor.bak.xlsx").is_file()


def test_creates_new_passage_in_the_same_locations_sheet():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    document = load_passage_document(path)
    result = save_passage(path, "Nowe", [
        {"lat": -10, "lon": 20, "locationType": "Przejscie"},
        {"lat": -11, "lon": 21, "locationType": "Przejscie"},
    ], document["revision"])
    created = passage(result, "Nowe")
    assert created["status"] == "development"
    assert [point["name"] for point in created["points"]] == ["Nowe 01", "Nowe 02"]
    assert [point["order"] for point in created["points"]] == [1, 2]


def test_moves_point_coordinates():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    points = passage(load_passage_document(path), "Alfa")["points"]
    points[0]["lat"], points[0]["lon"] = 1.25, 10.75
    result = save(path, "Alfa", points)
    assert passage(result, "Alfa")["points"][0]["lat"] == 1.25
    assert passage(result, "Alfa")["points"][0]["lon"] == 10.75


def test_changes_status_without_changing_passage_points():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    document = load_passage_document(path)
    alfa_before = passage(document, "Alfa")
    result = save_passage(
        path, "Alfa", alfa_before["points"], document["revision"],
        "Alfa", "development",
    )
    alfa_after = passage(result, "Alfa")
    assert alfa_after["status"] == "development"
    assert alfa_after["points"] == alfa_before["points"]


def test_accepts_only_the_deterministic_excel_float_serialization_difference():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    points = passage(load_passage_document(path), "Alfa")["points"]
    requested = -54.863456789012346
    points[0]["lat"] = requested
    result = save(path, "Alfa", points)
    saved = passage(result, "Alfa")["points"][0]["lat"]
    assert saved == float(f"{requested:.16g}")
    assert saved != requested


def test_detailed_verification_rejects_a_real_coordinate_difference():
    expected = [{
        "name": "Beagle 03", "order": 3, "lat": -54.865,
        "lon": -68.05522, "country": "", "locationType": "Przejscie",
        "notes": "",
    }]
    actual = [{**expected[0], "lon": -68.0552}]
    differences = _verification_differences(expected, actual)
    assert len(differences) == 1
    assert "Beagle 03: Lon" in differences[0]
    assert "-68.05522" in differences[0]
    assert "-68.0552" in differences[0]


def test_detailed_verification_reports_przejscie_lp_sequence():
    expected = [{
        "name": "A 01", "order": 1, "lat": 1.0, "lon": 2.0,
        "country": "", "locationType": "", "notes": "",
    }, {
        "name": "A 02", "order": 2, "lat": 3.0, "lon": 4.0,
        "country": "", "locationType": "", "notes": "",
    }]
    actual = [{**expected[0]}, {**expected[1], "order": 3}]
    differences = _verification_differences(expected, actual)
    assert differences[0] == "Przejscie_lp: oczekiwano [1, 2], odczytano [1, 3]"


def test_removes_point_and_renumbers_remaining_points():
    path = workdir() / "rejsy.xlsx"
    rows = standard_rows()
    rows.insert(3, ["Alfa 03", "Test", 2.5, 25.0, "Przejscie", "trzecie", "Alfa", 3])
    make_workbook(path, rows)
    points = passage(load_passage_document(path), "Alfa")["points"]
    del points[1]
    result = save(path, "Alfa", points)
    alfa = passage(result, "Alfa")
    assert [point["name"] for point in alfa["points"]] == ["Alfa 01", "Alfa 02"]
    assert [point["lat"] for point in alfa["points"]] == [1.0, 2.5]


def test_reorders_points_and_rebuilds_names_and_orders():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    points = list(reversed(passage(load_passage_document(path), "Alfa")["points"]))
    result = save(path, "Alfa", points)
    alfa = passage(result, "Alfa")
    assert [point["lat"] for point in alfa["points"]] == [2.0, 1.0]
    assert [point["name"] for point in alfa["points"]] == ["Alfa 01", "Alfa 02"]
    assert [point["order"] for point in alfa["points"]] == [1, 2]


def test_save_preserves_other_locations_passages_and_sheets():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    before = location_values(path)
    points = passage(load_passage_document(path), "Alfa")["points"]
    points[0]["lat"] = 1.5
    save(path, "Alfa", points)
    after = location_values(path)
    untouched_before = {row[0]: row for row in before if row[6] != "Alfa"}
    untouched_after = {row[0]: row for row in after if row[6] != "Alfa"}
    assert untouched_after == untouched_before
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        assert workbook["Pozostale"]["A1"].value == "nie zmieniaj"
    finally:
        workbook.close()


def test_cancel_read_only_path_does_not_change_excel():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    before = hashlib.sha256(path.read_bytes()).hexdigest()
    load_passage_document(path)
    load_passage_document(path)
    assert hashlib.sha256(path.read_bytes()).hexdigest() == before


@pytest.mark.parametrize("mutation,fragment", [
    (lambda points: points[:1], "co najmniej 2"),
    (lambda points: [{**points[0], "lat": 91}, points[1]], "zakres"),
    (lambda points: [{**points[0], "lon": "x"}, points[1]], "liczbą"),
])
def test_rejects_invalid_edited_data_without_changing_workbook(mutation, fragment):
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    document = load_passage_document(path)
    before = path.read_bytes()
    with pytest.raises(ValueError, match=fragment):
        save_passage(path, "Alfa", mutation(passage(document, "Alfa")["points"]), document["revision"], "Alfa")
    assert path.read_bytes() == before


def test_rejects_duplicate_or_gapped_existing_przejscie_lp():
    path = workdir() / "rejsy.xlsx"
    rows = standard_rows()
    rows[1][-1] = 1
    make_workbook(path, rows)
    with pytest.raises(ValueError, match="ciąg 1..2"):
        load_passage_document(path)


def test_rejects_stale_revision_without_changing_workbook():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    document = load_passage_document(path)
    before = path.read_bytes()
    with pytest.raises(RuntimeError, match="zmienił się"):
        save_passage(path, "Alfa", passage(document, "Alfa")["points"], "0" * 64, "Alfa")
    assert path.read_bytes() == before


def test_current_beagle_open_and_save_is_byte_for_byte_noop_on_copy():
    directory = workdir()
    path = directory / "rejsy.xlsx"
    shutil.copy2(ROOT / "routes" / "rejsy.xlsx", path)
    document = load_passage_document(path)
    beagle = passage(document, "Beagle")
    assert beagle["status"] == "development"
    before = path.read_bytes()
    result = save_passage(path, "Beagle", beagle["points"], document["revision"], "Beagle")
    assert path.read_bytes() == before
    assert result["revision"] == document["revision"]
    assert not path.with_name("rejsy.editor.bak.xlsx").exists()


def test_current_beagle_moves_two_existing_points_without_changing_identity_or_order():
    directory = workdir()
    path = directory / "rejsy.xlsx"
    shutil.copy2(ROOT / "routes" / "rejsy.xlsx", path)
    document = load_passage_document(path)
    beagle = passage(document, "Beagle")
    points = [dict(point) for point in beagle["points"]]
    original_identity = [(point["name"], point["order"]) for point in points]
    points[0]["lat"], points[0]["lon"] = (
        -54.91056698961305, -67.66891479492189,
    )
    points[2]["lat"], points[2]["lon"] = (
        -54.854477551567086, -67.86804199218751,
    )
    result = save_passage(
        path, "Beagle", points, document["revision"], "Beagle"
    )
    reread = passage(result, "Beagle")["points"]
    assert [(point["name"], point["order"]) for point in reread] == original_identity
    assert reread[0]["lat"] == float(f"{points[0]['lat']:.16g}")
    assert reread[0]["lon"] == float(f"{points[0]['lon']:.16g}")
    assert reread[2]["lat"] == float(f"{points[2]['lat']:.16g}")
    assert reread[2]["lon"] == float(f"{points[2]['lon']:.16g}")


def test_failed_temp_verification_reports_the_first_point_and_field(monkeypatch):
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    document = load_passage_document(path)
    points = [dict(point) for point in passage(document, "Alfa")["points"]]
    points[0]["lat"] = 1.25
    real_read = passage_editor._read_document

    def read_with_corrupted_temp(candidate: Path):
        result = real_read(candidate)
        if ".editor-" in candidate.name:
            passage(result, "Alfa")["points"][0]["lon"] = 10.5
        return result

    monkeypatch.setattr(passage_editor, "_read_document", read_with_corrupted_temp)
    with pytest.raises(
        RuntimeError,
        match=r"Pierwsza różnica: Alfa 01: Lon .*oczekiwano 10\.0.*odczytano 10\.5",
    ):
        save_passage(
            path, "Alfa", points, document["revision"], "Alfa"
        )


def test_local_http_editor_serves_assets_and_workbook_data():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path, standard_rows())
    server = create_editor_server(path, ROOT, port=0)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        base = f"http://127.0.0.1:{server.server_port}"
        with urlopen(base + "/", timeout=5) as response:
            assert "Edytor przejść" in response.read().decode("utf-8")
        with urlopen(base + "/api/passages", timeout=5) as response:
            payload = json.load(response)
        assert [item["name"] for item in payload["passages"]] == ["Alfa", "Beta"]
        assert payload["editorProtocol"] == EDITOR_PROTOCOL_VERSION
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)
