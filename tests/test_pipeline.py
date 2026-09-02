import json
import unittest
from contextlib import contextmanager
from pathlib import Path
from uuid import uuid4
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from rejsy_morskie.geocoding import CachedGeocoder, GeocodingCandidate
from rejsy_morskie.pipeline import generate_routes
from rejsy_morskie.sea_router import RouteResult


class FakeProvider:
    def __init__(self):
        self.calls = []

    def search(self, port, country):
        self.calls.append(port)
        coordinates = {"A": (42.0, 18.0), "B": (40.0, 16.0)}
        lat, lon = coordinates[port]
        return [GeocodingCandidate(port, lat, lon, "test")]


class FakeRouter:
    def __init__(self):
        self.penalties = []

    def route(self, start_lat, start_lon, end_lat, end_lon, *, penalty=None):
        self.penalties.append(penalty)
        return RouteResult(
            geometry={
                "type": "LineString",
                "coordinates": [[start_lon, start_lat], [end_lon, end_lat]],
            },
            distance_nm=123.45,
        )


@contextmanager
def work_directory():
    path = Path(__file__).parent / "_work" / uuid4().hex
    path.mkdir(parents=True)
    yield path


class PipelineTests(unittest.TestCase):
    def test_generate_routes_writes_all_outputs(self) -> None:
        with work_directory() as root:
            input_path = root / "input.xlsx"
            output_dir = root / "outputs"
            self._make_workbook(input_path)
            geocoder = CachedGeocoder(FakeProvider(), root / "cache.json")
            router = FakeRouter()

            outputs = generate_routes(input_path, output_dir, geocoder, router)

            workbook_path = input_path
            workbook = load_workbook(workbook_path, data_only=True)
            porty = workbook["Porty"]
            etapy = workbook["Etapy"]

            self.assertEqual(porty["G2"].value, 42)
            self.assertEqual(porty["H3"].value, 16)
            self.assertEqual(porty["E2"].value, "0")
            self.assertEqual(porty["E3"].value, "+1")
            self.assertEqual(porty["E2"].number_format, "@")
            self.assertEqual(porty["E3"].number_format, "@")
            self.assertEqual(etapy["N2"].value, "gotowy")
            self.assertEqual(etapy["F2"].value, 1)
            self.assertEqual(etapy["G2"].value, 2)
            self.assertEqual(etapy["J2"].value, "Dni 1–2")
            self.assertEqual(etapy["L2"].value, 123.5)
            self.assertEqual(
                etapy["M2"].value, str(Path("R1") / "geojson" / "01-A-B.geojson")
            )
            self.assertEqual(etapy.tables["EtapyTable"].ref, "A1:O2")
            self.assertTrue((output_dir / "R1" / "trasa.kml").exists())
            self.assertTrue((output_dir / "R1" / "geojson" / "01-A-B.geojson").exists())
            self.assertEqual(outputs[0], workbook_path)
            self.assertTrue((root / "input.bak.xlsx").exists())
            self.assertEqual(router.penalties, [8.0])

            kml = ElementTree.parse(output_dir / "R1" / "trasa.kml")
            namespace = {"kml": "http://www.opengis.net/kml/2.2"}
            line_coordinates = kml.find(
                ".//kml:Folder[kml:name='Etapy']//kml:LineString/kml:coordinates",
                namespace,
            )
            self.assertIsNotNone(line_coordinates)
            self.assertEqual(line_coordinates.text, "18.0,42.0,0 16.0,40.0,0")

    def test_route_points_force_geometry_but_create_one_logical_leg(self) -> None:
        with work_directory() as root:
            input_path = root / "input.xlsx"
            output_dir = root / "outputs"
            self._make_waypoint_workbook(input_path)
            provider = FakeProvider()
            geocoder = CachedGeocoder(provider, root / "cache.json")
            router = FakeRouter()

            generate_routes(input_path, output_dir, geocoder, router)

            workbook = load_workbook(input_path, data_only=True)
            porty = workbook["Porty"]
            etapy = workbook["Etapy"]
            self.assertEqual(provider.calls, ["A", "B"])
            self.assertEqual((porty["G3"].value, porty["H3"].value), (41, 17))
            self.assertEqual((porty["G4"].value, porty["H4"].value), (40.5, 16.5))
            self.assertEqual(etapy["E2"].value, "A → B")
            self.assertIsNone(etapy["A3"].value)
            self.assertEqual(etapy.tables["EtapyTable"].ref, "A1:O2")
            self.assertEqual(router.penalties, [8.0, 8.0, 8.0])

            geojson_path = output_dir / "R1" / "geojson" / "01-A-B.geojson"
            feature = json.loads(geojson_path.read_text(encoding="utf-8"))
            self.assertEqual(
                feature["geometry"]["coordinates"],
                [[18.0, 42.0], [17.0, 41.0], [16.5, 40.5], [16.0, 40.0]],
            )
            self.assertEqual(
                [item["name"] for item in feature["properties"]["route_points"]],
                ["Horn1", "Horn2"],
            )

            kml = ElementTree.parse(output_dir / "R1" / "trasa.kml")
            namespace = {"kml": "http://www.opengis.net/kml/2.2"}
            point_names = [
                item.text for item in kml.findall(
                    ".//kml:Folder[kml:name='Punkty trasy']/kml:Placemark/kml:name",
                    namespace,
                )
            ]
            self.assertEqual(point_names, ["Horn1"])

    @staticmethod
    def _make_workbook(path: Path) -> None:
        workbook = Workbook()
        rejsy = workbook.active
        rejsy.title = "Rejsy"
        rejsy.append(
            ["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"]
        )
        rejsy.append(["R1", "Test", "2024-12-07", "#0057B8", 8, None])

        porty = workbook.create_sheet("Porty")
        porty.append(
            [
                "Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy",
                "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ",
            ]
        )
        porty.append(["R1", 1, "A", None, "0", 0, None, None, None, None])
        porty.append([None, 2, "B", None, "+1", 1, None, None, None, None])
        porty["E2"].number_format = "@"
        porty["E3"].number_format = "@"

        etapy = workbook.create_sheet("Etapy")
        etapy.append(
            [
                "Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu",
                "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni",
                "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi",
            ]
        )
        etapy.append([
            "STARE", 99, "Nieaktualny", "Wynik", "Stary etap",
            99, 100, "2099-01-01", "2099-01-02", "Dni 99–100",
            "stare daty", 999, "stara.geojson", "stary", "nie używać",
        ])
        table = Table(displayName="EtapyTable", ref="A1:O2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        etapy.add_table(table)
        workbook.save(path)

    @staticmethod
    def _make_waypoint_workbook(path: Path) -> None:
        workbook = Workbook()
        rejsy = workbook.active
        rejsy.title = "Rejsy"
        rejsy.append(["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"])
        rejsy.append(["R1", "Test", "2024-12-07", "#0057B8", 8, None])
        porty = workbook.create_sheet("Porty")
        porty.append(["Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy", "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ"])
        porty.append(["R1", 1, "A", None, "0", 0, None, None, None, None])
        porty.append([None, 2, "Horn1", None, "+0", 0, 41.0, 17.0, None, "Punkt_trasy"])
        porty.append([None, 3, "Horn2", None, "+0", 0, 40.5, 16.5, None, "Punkt_trasy_ukryty"])
        porty.append([None, 4, "B", None, "3", 1, None, None, None, None])
        for row in range(2, 6):
            porty.cell(row, 5).number_format = "@"
        etapy = workbook.create_sheet("Etapy")
        etapy.append([
            "Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu",
            "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni",
            "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi",
        ])
        etapy.append([None] * 15)
        table = Table(displayName="EtapyTable", ref="A1:O2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        etapy.add_table(table)
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()
