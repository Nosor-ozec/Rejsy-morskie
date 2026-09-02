import unittest
from contextlib import contextmanager
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from rejsy_morskie.excel_io import load_input
from rejsy_morskie.geocoding import CachedGeocoder, GeocodingCandidate
from rejsy_morskie.pipeline import generate_routes
from rejsy_morskie.sea_router import RouteResult


class RecordingProvider:
    def __init__(self) -> None:
        self.calls: list[str] = []

    def search(self, port, country):
        self.calls.append(port)
        return [GeocodingCandidate(port, 50.0, 20.0, "test")]


class SimpleRouter:
    def __init__(self) -> None:
        self.calls: list[tuple[float, float, float, float]] = []

    def route(self, start_lat, start_lon, end_lat, end_lon, *, penalty=None):
        self.calls.append((start_lat, start_lon, end_lat, end_lon))
        return RouteResult(
            geometry={
                "type": "LineString",
                "coordinates": [[start_lon, start_lat], [end_lon, end_lat]],
            },
            distance_nm=10.0,
        )


@contextmanager
def work_directory():
    path = Path(__file__).parent / "_work" / uuid4().hex
    path.mkdir(parents=True)
    yield path


class LocationsTests(unittest.TestCase):
    def test_regular_port_uses_case_insensitive_location_without_geocoding(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            self._make_workbook(
                path,
                first_port="PUERTO TEST",
                first_coordinates=(None, None),
                locations=[["  puerto test  ", "Chile", -45.5, -72.8, "Port", None]],
            )
            provider = RecordingProvider()

            generate_routes(
                path,
                root / "outputs",
                CachedGeocoder(provider, root / "cache.json"),
                SimpleRouter(),
            )

            saved = load_workbook(path, data_only=True)["Porty"]
            self.assertEqual((saved["G2"].value, saved["H2"].value), (-45.5, -72.8))
            self.assertEqual(provider.calls, [])

    def test_manual_port_coordinates_take_priority_over_location(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            self._make_workbook(
                path,
                first_port="A",
                first_coordinates=(1.25, 2.5),
                locations=[["A", None, 50.0, 60.0, "Port", None]],
            )
            provider = RecordingProvider()

            generate_routes(
                path,
                root / "outputs",
                CachedGeocoder(provider, root / "cache.json"),
                SimpleRouter(),
            )

            saved = load_workbook(path, data_only=True)["Porty"]
            self.assertEqual((saved["G2"].value, saved["H2"].value), (1.25, 2.5))
            self.assertEqual(provider.calls, [])

    def test_route_point_uses_location_and_is_never_geocoded(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            self._make_workbook(
                path,
                route_point="HORN1",
                locations=[[" horn1 ", None, -56.1, -67.3, "Punkt_trasy", "ręczne"]],
            )
            provider = RecordingProvider()

            generate_routes(
                path,
                root / "outputs",
                CachedGeocoder(provider, root / "cache.json"),
                SimpleRouter(),
            )

            saved = load_workbook(path, data_only=True)["Porty"]
            self.assertEqual((saved["G3"].value, saved["H3"].value), (-56.1, -67.3))
            self.assertEqual(provider.calls, [])

    def test_kanal_beagle_1_uses_locations_before_route_point_validation(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            self._make_workbook(
                path,
                route_point="Kanal Beagle 1",
                locations=[
                    [
                        "Kanal Beagle 1",
                        None,
                        -54.85415,
                        -68.05522,
                        "Punkt_trasy",
                        "ręcznie zatwierdzone",
                    ]
                ],
            )
            before_locations = list(
                load_workbook(path, data_only=False)["Lokalizacje"].iter_rows(
                    values_only=True
                )
            )
            provider = RecordingProvider()
            router = SimpleRouter()

            _, calls = load_input(path)
            point = next(call for call in calls if call.port == "Kanal Beagle 1")
            self.assertEqual((point.lat, point.lon), (-54.85415, -68.05522))
            self.assertEqual(point.coordinates_source, "lokalizacje")

            generate_routes(
                path,
                root / "outputs",
                CachedGeocoder(provider, root / "cache.json"),
                router,
            )

            workbook = load_workbook(path, data_only=True)
            self.assertEqual(
                (workbook["Porty"]["G3"].value, workbook["Porty"]["H3"].value),
                (-54.85415, -68.05522),
            )
            self.assertEqual(provider.calls, [])
            self.assertEqual(len(router.calls), 2)
            self.assertEqual(router.calls[0][2:], (-54.85415, -68.05522))
            after_locations = list(
                workbook["Lokalizacje"].iter_rows(values_only=True)
            )
            self.assertEqual(after_locations, before_locations)

    def test_przyladek_horn_uses_locations_before_route_point_validation(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            self._make_workbook(
                path,
                route_point="Przyladek Horn",
                locations=[
                    [
                        "Przyladek Horn",
                        "Argentyna",
                        -56.0633,
                        -67.2892,
                        "Punkt_trasy",
                        "ręcznie zatwierdzone",
                    ]
                ],
            )
            before_locations = list(
                load_workbook(path, data_only=False)["Lokalizacje"].iter_rows(
                    values_only=True
                )
            )
            provider = RecordingProvider()
            router = SimpleRouter()

            _, calls = load_input(path)
            point = next(call for call in calls if call.port == "Przyladek Horn")
            self.assertEqual((point.lat, point.lon), (-56.0633, -67.2892))
            self.assertEqual(point.coordinates_source, "lokalizacje")

            generate_routes(
                path,
                root / "outputs",
                CachedGeocoder(provider, root / "cache.json"),
                router,
            )

            workbook = load_workbook(path, data_only=True)
            self.assertEqual(
                (workbook["Porty"]["G3"].value, workbook["Porty"]["H3"].value),
                (-56.0633, -67.2892),
            )
            self.assertEqual(provider.calls, [])
            self.assertEqual(len(router.calls), 2)
            self.assertEqual(router.calls[0][2:], (-56.0633, -67.2892))
            after_locations = list(
                workbook["Lokalizacje"].iter_rows(values_only=True)
            )
            self.assertEqual(after_locations, before_locations)

    def test_unknown_route_point_reports_error_before_geocoding(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            self._make_workbook(path, route_point="Nieznany", locations=[])
            provider = RecordingProvider()

            with self.assertRaisesRegex(
                ValueError, "brak Lat/Lon w Porty i wpisu w Lokalizacje"
            ):
                generate_routes(
                    path,
                    root / "outputs",
                    CachedGeocoder(provider, root / "cache.json"),
                    SimpleRouter(),
                )
            self.assertEqual(provider.calls, [])

    def test_duplicate_normalized_location_name_is_rejected(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            self._make_workbook(
                path,
                locations=[
                    ["Test", None, 1.0, 2.0, "Port", None],
                    [" test ", None, 3.0, 4.0, "Kotwicowisko", None],
                ],
            )

            with self.assertRaisesRegex(ValueError, "powtórzona Nazwa po normalizacji"):
                load_input(path)

    def test_locations_values_remain_unchanged_after_generation(self):
        with work_directory() as root:
            path = root / "input.xlsx"
            locations = [
                ["A", "Test", 1.5, 2.5, "Kotwicowisko", "zatwierdzone"],
                ["Inne", None, -10.0, 120.0, "Wyspa", None],
            ]
            self._make_workbook(
                path,
                first_coordinates=(None, None),
                locations=locations,
            )
            before = list(
                load_workbook(path, data_only=False)["Lokalizacje"].iter_rows(
                    values_only=True
                )
            )

            generate_routes(
                path,
                root / "outputs",
                CachedGeocoder(RecordingProvider(), root / "cache.json"),
                SimpleRouter(),
            )

            after = list(
                load_workbook(path, data_only=False)["Lokalizacje"].iter_rows(
                    values_only=True
                )
            )
            self.assertEqual(after, before)

    @staticmethod
    def _make_workbook(
        path: Path,
        *,
        first_port: str = "A",
        first_coordinates=(10.0, 20.0),
        route_point: str | None = None,
        locations=None,
    ) -> None:
        workbook = Workbook()
        rejsy = workbook.active
        rejsy.title = "Rejsy"
        rejsy.append(
            ["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"]
        )
        rejsy.append(["R1", "Test", "2024-12-07", "#0057B8", 8, None])

        porty = workbook.create_sheet("Porty")
        porty.append(
            [
                "Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy",
                "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ",
            ]
        )
        porty.append(
            ["R1", 1, first_port, None, "0", 0, first_coordinates[0], first_coordinates[1], None, None]
        )
        end_order = 2
        if route_point is not None:
            porty.append([None, 2, route_point, None, "+0", 0, None, None, None, "Punkt_trasy"])
            end_order = 3
        porty.append([None, end_order, "B", None, "+1", 1, 11.0, 21.0, None, None])
        for row in range(2, porty.max_row + 1):
            porty.cell(row, 5).number_format = "@"

        etapy = workbook.create_sheet("Etapy")
        etapy.append(
            [
                "Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu",
                "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni",
                "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi",
            ]
        )
        etapy.append([None] * 15)
        table = Table(displayName="EtapyTable", ref="A1:O2")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        etapy.add_table(table)

        lokalizacje = workbook.create_sheet("Lokalizacje")
        lokalizacje.append(["Nazwa", "Kraj", "Lat", "Lon", "Typ", "Uwagi"])
        for row in locations or []:
            lokalizacje.append(row)
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()
