import json
import unittest
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from xml.etree import ElementTree

from rejsy_morskie.geocoding import CachedGeocoder, GeocodingCandidate
from rejsy_morskie.pipeline import generate_routes
from rejsy_morskie.sea_router import RouteResult
from rejsy_morskie.web import build_local_site, publish_site


class RecordingProvider:
    def __init__(self):
        self.calls = []

    def search(self, port, country):
        self.calls.append(port)
        return [GeocodingCandidate(port, 50.0, 20.0, "test")]


class StraightRouter:
    def route(self, start_lat, start_lon, end_lat, end_lon, *, penalty=None):
        return RouteResult(
            geometry={
                "type": "LineString",
                "coordinates": [[start_lon, start_lat], [end_lon, end_lat]],
            },
            distance_nm=100.0,
        )


class WebTests(unittest.TestCase):
    def test_local_site_and_publication_use_the_same_files(self) -> None:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        rejsy_path = root / "rejsy.xlsx"
        media_path = root / "media.xlsx"
        outputs = root / "outputs"
        assets = root / "assets"
        site = root / "site"
        docs = root / "docs"
        self._make_rejsy(rejsy_path)
        self._make_media(media_path)
        self._make_geojson(outputs)
        self._make_assets(assets)

        build_local_site(rejsy_path, media_path, outputs, site, assets)
        route = json.loads((site / "data" / "route.json").read_text(encoding="utf-8"))
        media = json.loads((site / "data" / "media.json").read_text(encoding="utf-8"))

        self.assertEqual([item["name"] for item in route["ports"]], ["A", "B"])
        self.assertEqual([item["name"] for item in route["routePoints"]], ["Horn1"])
        self.assertEqual(route["legs"][0]["coordinates"], [[0.0, 0.0], [1.0, 0.0], [2.5, 0.0], [4.0, 0.0]])
        self.assertEqual(len(media["media"]), 6)
        self.assertFalse(media["media"][0]["atSea"])
        self.assertTrue(media["media"][1]["atSea"])
        self.assertAlmostEqual(media["media"][1]["position"][0], 1.5, places=2)
        self.assertEqual(media["media"][1]["description"], "Opis na morzu")
        self.assertFalse(media["media"][2]["atSea"])
        self.assertEqual(media["media"][2]["baseType"], "Punkt_trasy")
        self.assertAlmostEqual(media["media"][2]["position"][0], 1.0, places=2)
        self.assertTrue(media["media"][3]["atSea"])
        self.assertAlmostEqual(media["media"][3]["position"][0], 1.25, places=2)
        self.assertTrue(media["media"][4]["atSea"])
        self.assertAlmostEqual(media["media"][4]["position"][0], 2.5, places=2)
        self.assertTrue(media["media"][5]["atSea"])
        self.assertAlmostEqual(media["media"][5]["position"][0], 3.2, places=2)

        docs.mkdir()
        (docs / "README.md").write_text("zachowaj", encoding="utf-8")
        publish_site(site, docs)
        self.assertEqual((docs / "README.md").read_text(encoding="utf-8"), "zachowaj")
        for relative in ("index.html", "app.js", "style.css", "data/route.json", "data/media.json"):
            self.assertEqual((site / relative).read_bytes(), (docs / relative).read_bytes())

    def test_repeated_port_visits_work_through_routing_kml_and_leaflet(self) -> None:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        rejsy_path = root / "rejsy.xlsx"
        media_path = root / "media.xlsx"
        outputs = root / "outputs"
        assets = root / "assets"
        site = root / "site"
        self._make_repeated_port_rejsy(rejsy_path)
        self._make_empty_media(media_path)
        self._make_assets(assets)
        provider = RecordingProvider()

        generate_routes(
            rejsy_path,
            outputs,
            CachedGeocoder(provider, root / "cache.json"),
            StraightRouter(),
        )
        build_local_site(rejsy_path, media_path, outputs, site, assets)

        workbook = load_workbook(rejsy_path, data_only=True)
        porty = workbook["Porty"]
        self.assertEqual((porty["G3"].value, porty["H3"].value), (41.39, 2.17))
        self.assertEqual((porty["G5"].value, porty["H5"].value), (41.39, 2.17))
        self.assertEqual(provider.calls, [])
        etapy = workbook["Etapy"]
        self.assertEqual(
            [etapy.cell(row, 5).value for row in range(2, 6)],
            [
                "Triest → Barcelona",
                "Barcelona → Split",
                "Split → Barcelona",
                "Barcelona → Triest",
            ],
        )

        kml = ElementTree.parse(outputs / "R1" / "trasa.kml")
        namespace = {"kml": "http://www.opengis.net/kml/2.2"}
        names = [
            item.text
            for item in kml.findall(
                ".//kml:Folder[kml:name='Porty']/kml:Placemark/kml:name",
                namespace,
            )
        ]
        self.assertEqual(names.count("Barcelona"), 2)

        route = json.loads((site / "data" / "route.json").read_text(encoding="utf-8"))
        self.assertEqual([item["name"] for item in route["ports"]].count("Barcelona"), 2)
        self.assertEqual(
            [item["visitId"] for item in route["ports"]],
            ["R1:1", "R1:2", "R1:3", "R1:4", "R1:5"],
        )
        self.assertEqual(len(route["legs"]), 4)
        self.assertTrue((site / "index.html").exists())

    def test_media_by_repeated_port_name_is_available_at_every_visit(self) -> None:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        rejsy_path = root / "rejsy.xlsx"
        media_path = root / "media.xlsx"
        outputs = root / "outputs"
        assets = root / "assets"
        self._make_repeated_port_rejsy(rejsy_path)
        self._make_media_rows(
            media_path,
            [["Barcelona_1", None, None, 0, "Barcelona", "https://example.com/b", "TAK"]],
        )
        self._make_assets(assets)
        generate_routes(
            rejsy_path,
            outputs,
            CachedGeocoder(RecordingProvider(), root / "cache.json"),
            StraightRouter(),
        )

        site = root / "site"
        build_local_site(rejsy_path, media_path, outputs, site, assets)
        media = json.loads((site / "data" / "media.json").read_text(encoding="utf-8"))["media"]
        self.assertEqual([item["baseVisitId"] for item in media], ["R1:2", "R1:4"])
        self.assertEqual([item["id"] for item in media], ["Barcelona_1", "Barcelona_1"])

    def test_legacy_visit_order_is_ignored_for_repeated_port(self) -> None:
        root, rejsy_path, media_path, outputs, assets = self._repeated_media_case(
            [["Barcelona_1", None, None, 0, "Druga Barcelona", "https://example.com/b", "TAK", 4]]
        )

        site = root / "site"
        build_local_site(rejsy_path, media_path, outputs, site, assets)
        media = json.loads((site / "data" / "media.json").read_text(encoding="utf-8"))["media"]
        self.assertEqual([item["baseVisitId"] for item in media], ["R1:2", "R1:4"])
        self.assertEqual([item["position"] for item in media], [[41.39, 2.17], [41.39, 2.17]])

    def test_legacy_visit_order_does_not_bind_media_to_another_name(self) -> None:
        root, rejsy_path, media_path, outputs, assets = self._repeated_media_case(
            [["Barcelona_1", None, None, 0, "Błędny port", "https://example.com/b", "TAK", 3]]
        )

        site = root / "site"
        build_local_site(rejsy_path, media_path, outputs, site, assets)
        media = json.loads((site / "data" / "media.json").read_text(encoding="utf-8"))["media"]
        self.assertEqual([item["baseVisitId"] for item in media], ["R1:2", "R1:4"])

    def test_legacy_nonexistent_visit_order_is_ignored(self) -> None:
        root, rejsy_path, media_path, outputs, assets = self._repeated_media_case(
            [["Barcelona_1", None, None, 0, "Brak wizyty", "https://example.com/b", "TAK", 99]]
        )

        site = root / "site"
        build_local_site(rejsy_path, media_path, outputs, site, assets)
        media = json.loads((site / "data" / "media.json").read_text(encoding="utf-8"))["media"]
        self.assertEqual([item["baseVisitId"] for item in media], ["R1:2", "R1:4"])

    def test_repeated_port_visits_can_have_different_stay_marker_data(self) -> None:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        rejsy_path = root / "rejsy.xlsx"
        media_path = root / "media.xlsx"
        outputs = root / "outputs"
        assets = root / "assets"
        site = root / "site"
        self._make_repeated_port_rejsy(rejsy_path)
        workbook = load_workbook(rejsy_path)
        workbook["Porty"].cell(5, 6).value = 2
        workbook.save(rejsy_path)
        self._make_empty_media(media_path)
        self._make_assets(assets)
        generate_routes(
            rejsy_path, outputs,
            CachedGeocoder(RecordingProvider(), root / "cache.json"),
            StraightRouter(),
        )
        build_local_site(rejsy_path, media_path, outputs, site, assets)
        route = json.loads((site / "data" / "route.json").read_text(encoding="utf-8"))
        barcelona = [item for item in route["ports"] if item["name"] == "Barcelona"]
        self.assertEqual([item["stayDays"] for item in barcelona], [1, 2])
        app = (Path(__file__).resolve().parents[1] / "docs" / "app.js").read_text(encoding="utf-8")
        self.assertIn("if (port.stayDays > 1) markerOptions.icon = longStayIcon()", app)
        self.assertIn("return L.icon({", app)
        self.assertIn("iconUrl: LONG_STAY_PIN_URL", app)
        self.assertIn("shadowUrl: 'vendor/leaflet/images/marker-shadow.png'", app)
        self.assertIn("iconSize: [25, 41]", app)
        self.assertNotIn("return L.divIcon({", app)
        self.assertIn("pane: 'port-markers'", app)
        self.assertIn("zIndexOffset: 1000", app)
        self.assertIn("longStayIcon", app)
        self.assertIn("Postój: ${stayDays} dni", app)

    def test_regular_port_marker_stays_above_technical_point_at_same_position(self) -> None:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        rejsy_path = root / "rejsy.xlsx"
        media_path = root / "media.xlsx"
        outputs = root / "outputs"
        assets = root / "assets"
        site = root / "site"
        self._make_rejsy(rejsy_path)
        workbook = load_workbook(rejsy_path)
        workbook["Porty"].cell(3, 7).value = 0
        workbook["Porty"].cell(3, 8).value = 0
        workbook.save(rejsy_path)
        self._make_empty_media(media_path)
        self._make_geojson(outputs)
        self._make_assets(assets)

        build_local_site(rejsy_path, media_path, outputs, site, assets)
        route = json.loads((site / "data" / "route.json").read_text(encoding="utf-8"))
        port = next(item for item in route["ports"] if item["name"] == "A")
        point = next(item for item in route["routePoints"] if item["name"] == "Horn1")
        self.assertEqual(port["position"], point["position"])

        app = (Path(__file__).resolve().parents[1] / "docs" / "app.js").read_text(
            encoding="utf-8"
        )
        self.assertIn("map.createPane('port-markers')", app)
        self.assertIn("portPane.style.zIndex = '660'", app)
        self.assertIn("technicalPane.style.zIndex = '620'", app)
        self.assertIn("pane: 'port-markers'", app)
        self.assertIn("pane: 'technical-markers'", app)
        self.assertIn("zIndexOffset: 1000", app)

    def test_media_selector_supports_visible_and_hidden_route_points(self) -> None:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        rejsy_path = root / "rejsy.xlsx"
        media_path = root / "media.xlsx"
        outputs = root / "outputs"
        assets = root / "assets"
        site = root / "site"
        self._make_rejsy(rejsy_path)
        self._make_media_rows(
            media_path,
            [
                ["Horn1_1", None, None, 0, "Widoczny", "https://example.com/h1", "TAK", 2],
                ["Horn2_1", None, None, 0, "Ukryty", "https://example.com/h2", "TAK", 3],
                ["Horn2_2", None, None, 0.7, "Po ukrytym", "https://example.com/h2a", "TAK", 3],
            ],
            include_visit_order=True,
        )
        self._make_geojson(outputs)
        self._make_assets(assets)

        build_local_site(rejsy_path, media_path, outputs, site, assets)
        media = json.loads((site / "data" / "media.json").read_text(encoding="utf-8"))["media"]
        self.assertEqual([item["baseVisitId"] for item in media], ["R1:2", "R1:3", "R1:3"])
        self.assertFalse(media[0]["atSea"])
        self.assertTrue(media[1]["atSea"])
        self.assertTrue(media[2]["atSea"])
        self.assertAlmostEqual(media[2]["position"][0], 3.2, places=2)

    def _repeated_media_case(self, rows: list[list[object]]) -> tuple[Path, Path, Path, Path, Path]:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        rejsy_path = root / "rejsy.xlsx"
        media_path = root / "media.xlsx"
        outputs = root / "outputs"
        assets = root / "assets"
        self._make_repeated_port_rejsy(rejsy_path)
        self._make_media_rows(media_path, rows, include_visit_order=True)
        self._make_assets(assets)
        generate_routes(
            rejsy_path,
            outputs,
            CachedGeocoder(RecordingProvider(), root / "cache.json"),
            StraightRouter(),
        )
        return root, rejsy_path, media_path, outputs, assets

    @staticmethod
    def _make_rejsy(path: Path) -> None:
        workbook = Workbook()
        rejsy = workbook.active
        rejsy.title = "Rejsy"
        rejsy.append(["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"])
        rejsy.append(["R1", "Test", "2024-01-01", "Niebieski", 4, None])
        porty = workbook.create_sheet("Porty")
        porty.append(["Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy", "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ"])
        porty.append(["R1", 1, "A", None, "0", 0, 0, 0, None, None])
        porty.append([None, 2, "Horn1", None, "+0", 0, 1, 0, None, "Punkt_trasy"])
        porty.append([None, 3, "Horn2", None, "+0", 0, 2.5, 0, None, "Punkt_trasy_ukryty"])
        porty.append([None, 4, "B", None, "3", 0, 4, 0, None, None])
        for row in range(2, 6):
            porty.cell(row, 5).number_format = "@"
        etapy = workbook.create_sheet("Etapy")
        etapy.append(["Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu", "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni", "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi"])
        etapy.append(["R1", 1, "A", "B", "A → B", 1, 4, "2024-01-01", "2024-01-04", "Dni 1–4", "2024-01-01 – 2024-01-04", 180, str(Path("R1") / "geojson" / "01-A-B.geojson"), "gotowy", None])
        workbook.save(path)

    @staticmethod
    def _make_media(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Filmy"
        sheet.append(["Film_ID", "Typ", "Powiazanie", "Dzien_od_portu", "Opis", "URL_Google_Drive", "Aktywny"])
        sheet.append(["A_1", None, None, 0, "Opis portu", "https://example.com/port", "TAK"])
        sheet.append(["A_2", None, None, 1.5, "Opis na morzu", "https://example.com/morze", "TAK"])
        sheet.append(["Horn1_1", None, None, 0, "Na punkcie", "https://example.com/horn1", "TAK"])
        sheet.append(["Horn1_2", None, None, 0.25, "Po Horn1", "https://example.com/horn1-after", "TAK"])
        sheet.append(["Horn2_1", None, None, 0, "Na ukrytym punkcie", "https://example.com/horn2", "TAK"])
        sheet.append(["Horn2_2", None, None, 0.7, "Po Horn2", "https://example.com/horn2-after", "TAK"])
        sheet.append(["B_1", None, None, 0, "Nieaktywne", "https://example.com/off", "NIE"])
        workbook.save(path)

    @staticmethod
    def _make_empty_media(path: Path) -> None:
        WebTests._make_media_rows(path, [])

    @staticmethod
    def _make_media_rows(
        path: Path,
        rows: list[list[object]],
        *,
        include_visit_order: bool = False,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Filmy"
        headers = ["Film_ID", "Typ", "Powiazanie", "Dzien_od_portu", "Opis", "URL_Google_Drive", "Aktywny"]
        if include_visit_order:
            headers.append("Kolejnosc_wizyty")
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    @staticmethod
    def _make_repeated_port_rejsy(path: Path) -> None:
        workbook = Workbook()
        rejsy = workbook.active
        rejsy.title = "Rejsy"
        rejsy.append(["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"])
        rejsy.append(["R1", "Powtórzone porty", "2024-12-07", "Niebieski", 4, None])
        porty = workbook.create_sheet("Porty")
        porty.append(["Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy", "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ"])
        porty.append(["R1", 1, "Triest", "Włochy", "0", 0, 45.65, 13.78, None, None])
        porty.append([None, 2, "Barcelona", "Hiszpania", "+1", 1, None, None, None, None])
        porty.append([None, 3, "Split", "Chorwacja", "+1", 1, 43.51, 16.44, None, None])
        porty.append([None, 4, "Barcelona", "Hiszpania", "+1", 1, None, None, None, None])
        porty.append([None, 5, "Triest", "Włochy", "+1", 1, 45.65, 13.78, None, None])
        for row in range(2, 7):
            porty.cell(row, 5).number_format = "@"
        etapy = workbook.create_sheet("Etapy")
        etapy.append(["Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu", "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni", "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi"])
        lokalizacje = workbook.create_sheet("Lokalizacje")
        lokalizacje.append(["Nazwa", "Kraj", "Lat", "Lon", "Typ", "Uwagi"])
        lokalizacje.append(["Barcelona", "Hiszpania", 41.39, 2.17, "Port", None])
        workbook.save(path)

    @staticmethod
    def _make_geojson(outputs: Path) -> None:
        path = outputs / "R1" / "geojson" / "01-A-B.geojson"
        path.parent.mkdir(parents=True)
        path.write_text(json.dumps({"type": "Feature", "geometry": {"type": "LineString", "coordinates": [[0, 0], [0, 1], [0, 2.5], [0, 4]]}, "properties": {}}), encoding="utf-8")

    @staticmethod
    def _make_assets(assets: Path) -> None:
        assets.mkdir()
        for name in ("index.html", "app.js", "style.css"):
            (assets / name).write_text(name, encoding="utf-8")
        vendor = assets / "vendor"
        vendor.mkdir()
        (vendor / "leaflet.js").write_text("leaflet", encoding="utf-8")


if __name__ == "__main__":
    unittest.main()
