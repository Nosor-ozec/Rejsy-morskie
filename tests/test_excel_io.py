import unittest
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from rejsy_morskie.excel_io import _inherit_voyage_id, load_input, write_results
from rejsy_morskie.schedule import calculate_schedule


class ExcelIoTests(unittest.TestCase):
    def test_blank_voyage_id_inherits_previous_row(self) -> None:
        self.assertEqual(_inherit_voyage_id(None, "R1", 3), "R1")
        self.assertEqual(_inherit_voyage_id("", "R1", 4), "R1")
        self.assertEqual(_inherit_voyage_id("R2", "R1", 5), "R2")

    def test_first_port_requires_voyage_id(self) -> None:
        with self.assertRaisesRegex(ValueError, "pierwszy Rejs_ID nie może być pusty"):
            _inherit_voyage_id(None, None, 2)

    def test_calculation_preserves_exact_when_and_stay_values_in_workbook(self) -> None:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        path = root / "rejsy-copy.xlsx"
        workbook = Workbook()
        rejsy = workbook.active
        rejsy.title = "Rejsy"
        rejsy.append(["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"])
        rejsy.append(["R1", "Test", datetime(2024, 12, 7), "Niebieski", 4, None])
        porty = workbook.create_sheet("Porty")
        porty.append(["Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy", "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ"])
        input_values = ["0", "+1", "4", "2024-12-13"]
        stay_values = [2, 2, 2, 3]
        for order, (when, stay) in enumerate(
            zip(input_values, stay_values), start=1
        ):
            porty.append(["R1" if order == 1 else None, order, f"P{order}", None, when, stay, None, None, None, None])
            porty.cell(order + 1, 5).number_format = "@"
        etapy = workbook.create_sheet("Etapy")
        etapy.append(["Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu", "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni", "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi"])
        workbook.save(path)

        voyages, calls = load_input(path)
        legs = calculate_schedule(voyages["R1"], calls)
        write_results(path, path, calls, legs)

        saved = load_workbook(path, data_only=False)["Porty"]
        self.assertEqual([saved.cell(row, 5).value for row in range(2, 6)], input_values)
        self.assertEqual([saved.cell(row, 6).value for row in range(2, 6)], stay_values)
        self.assertTrue(all(saved.cell(row, 5).data_type == "s" for row in range(2, 6)))
        self.assertTrue(all(saved.cell(row, 5).number_format == "@" for row in range(2, 6)))
        self.assertTrue((root / "rejsy-copy.bak.xlsx").exists())

    def test_both_route_point_types_keep_manual_coordinates(self) -> None:
        for point_type in ("Punkt_trasy", "Punkt_trasy_ukryty"):
            with self.subTest(point_type=point_type):
                path = self._make_point_workbook(point_type=point_type)
                voyages, calls = load_input(path)
                point = calls[1]
                self.assertEqual((point.lat, point.lon), (11.25, 22.5))
                self.assertEqual(point.call_type, point_type)

                point.lat, point.lon = 99.0, 88.0
                legs = calculate_schedule(voyages["R1"], calls)
                write_results(path, path, calls, legs)

                saved = load_workbook(path, data_only=False)["Porty"]
                self.assertEqual((saved["G3"].value, saved["H3"].value), (11.25, 22.5))
                self.assertEqual(saved["E3"].value, "+0")
                self.assertEqual(saved["F3"].value, 0)
                self.assertEqual(saved["J3"].value, point_type)

    def test_route_point_requires_lat_and_lon(self) -> None:
        path = self._make_point_workbook(lat=None, lon=None)
        with self.assertRaisesRegex(ValueError, "Lat i Lon są obowiązkowe"):
            load_input(path)

    def test_route_point_requires_zero_stay(self) -> None:
        path = self._make_point_workbook(stay=1)
        with self.assertRaisesRegex(ValueError, "Postoj_dni musi wynosić 0"):
            load_input(path)

    @staticmethod
    def _make_point_workbook(
        *, point_type: str = "Punkt_trasy", lat=11.25, lon=22.5, stay: int = 0
    ) -> Path:
        root = Path(__file__).parent / "_work" / uuid4().hex
        root.mkdir(parents=True)
        path = root / "route-point.xlsx"
        workbook = Workbook()
        rejsy = workbook.active
        rejsy.title = "Rejsy"
        rejsy.append(["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"])
        rejsy.append(["R1", "Test", "2024-12-07", "Niebieski", 4, None])
        porty = workbook.create_sheet("Porty")
        porty.append(["Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy", "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ"])
        porty.append(["R1", 1, "A", None, "0", 0, 10.0, 20.0, None, None])
        porty.append([None, 2, "Horn1", None, "+0", stay, lat, lon, None, point_type])
        porty.append([None, 3, "B", None, "2", 1, 12.0, 24.0, None, None])
        for row in range(2, 5):
            porty.cell(row, 5).number_format = "@"
        etapy = workbook.create_sheet("Etapy")
        etapy.append(["Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu", "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni", "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi"])
        workbook.save(path)
        return path


if __name__ == "__main__":
    unittest.main()
