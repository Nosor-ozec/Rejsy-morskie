from __future__ import annotations

import hashlib
import json
import shutil
import threading
from pathlib import Path
from urllib.request import urlopen
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from rejsy_morskie.port_editor import (
    EDITOR_PROTOCOL_VERSION,
    create_port_editor_server,
    load_port_document,
    save_voyage_ports,
)


ROOT = Path(__file__).resolve().parents[1]


def test_windows_wrapper_is_ascii_crlf_and_uses_project_relative_launcher():
    raw = (ROOT / "Edytuj-Porty.cmd").read_bytes()

    assert not raw.startswith(b"\xef\xbb\xbf")
    text = raw.decode("ascii")
    assert raw.count(b"\n") == raw.count(b"\r\n")
    assert 'cd /d "%~dp0"' in text
    assert 'set "PYTHON=%~dp0.venv\\Scripts\\python.exe"' in text
    assert 'set "REJSY=%~dp0routes\\rejsy.xlsx"' in text
    assert (
        '"%PYTHON%" -m rejsy_morskie.cli edit-ports "%REJSY%" '
        '--project-root "." --host 127.0.0.1 --port 8767'
    ) in text


def workdir() -> Path:
    path = Path(__file__).resolve().parent / "_work" / uuid4().hex
    path.mkdir(parents=True)
    return path


def make_workbook(path: Path) -> None:
    workbook = Workbook()
    rejsy = workbook.active
    rejsy.title = "Rejsy"
    rejsy.append(["Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"])
    rejsy.append(["R1", "Test", "2024-01-01", "Niebieski", 4, None])
    porty = workbook.create_sheet("Porty")
    porty.append(["Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy", "Postoj_dni", "Lat", "Lon", "Uwagi", "Typ"])
    porty.append(["R1", 1, "Triest", "Włochy", "0", 0, 45.65, 13.78, None, None])
    porty.append([None, 2, "Barcelona", "Hiszpania", "+1", 1, None, None, None, None])
    porty.append([None, 3, "Punkt A", None, "+0", 0, None, None, "techniczny", "Punkt_trasy"])
    porty.append([None, 4, "Barcelona", "Hiszpania", "+1", 2, None, None, None, None])
    porty.append([None, 5, "Triest", "Włochy", "+2", 1, 45.65, 13.78, None, None])
    for row in range(2, 7):
        porty.cell(row, 5).number_format = "@"
    etapy = workbook.create_sheet("Etapy")
    etapy.append(["Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu", "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni", "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi"])
    etapy["A2"] = "wynik pozostaje"
    locations = workbook.create_sheet("Lokalizacje")
    locations.append(["Nazwa", "Kraj", "Lat", "Lon", "Typ", "Uwagi", "Przejscie", "Przejscie_lp", "Przejscie_status"])
    locations.append(["Barcelona", "Hiszpania", 41.39, 2.17, "Port", "wspólna", None, None, None])
    locations.append(["Punkt A", None, 41.8, 5.0, "Punkt_trasy", "zachowaj", None, None, None])
    workbook.save(path)


def voyage(document: dict[str, object]) -> dict[str, object]:
    return document["voyages"][0]


def save(path: Path, calls: list[dict[str, object]]) -> dict[str, object]:
    document = load_port_document(path)
    return save_voyage_ports(path, "R1", calls, document["revision"])


def copy_calls(path: Path) -> list[dict[str, object]]:
    return [dict(item) for item in voyage(load_port_document(path))["calls"]]


def location_row(path: Path, name: str) -> tuple[object, ...]:
    workbook = load_workbook(path, data_only=False)
    try:
        sheet = workbook["Lokalizacje"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip().casefold() == name.casefold():
                return row
        raise AssertionError(name)
    finally:
        workbook.close()


def test_moves_existing_port_and_persists_one_shared_location_for_both_visits():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    calls = copy_calls(path)
    for call in calls:
        if call["name"] == "Barcelona":
            call["lat"], call["lon"] = 42.0, 3.0
    result = save(path, calls)
    barcelona = [call for call in voyage(result)["calls"] if call["name"] == "Barcelona"]
    assert [(call["lat"], call["lon"]) for call in barcelona] == [(42.0, 3.0), (42.0, 3.0)]
    assert location_row(path, "Barcelona")[2:4] == (42.0, 3.0)
    assert path.with_name("rejsy.ports-editor.bak.xlsx").is_file()


def test_location_restores_both_visits_after_porty_coordinates_are_cleared():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    calls = copy_calls(path)
    for call in calls:
        if call["name"] == "Barcelona":
            call["lat"], call["lon"] = 42.0, 3.0
    save(path, calls)
    workbook = load_workbook(path)
    for row in (3, 5):
        workbook["Porty"].cell(row, 7).value = None
        workbook["Porty"].cell(row, 8).value = None
    workbook.save(path)
    restored = [call for call in voyage(load_port_document(path))["calls"] if call["name"] == "Barcelona"]
    assert [(call["lat"], call["lon"], call["coordinatesSource"]) for call in restored] == [
        (42.0, 3.0, "lokalizacje"), (42.0, 3.0, "lokalizacje")
    ]


def test_adds_port_and_renumbers_calls_without_touching_etapy():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    calls = copy_calls(path)
    calls.insert(1, {
        "sourceVisitId": "", "name": "Nowy Port", "country": "Polska",
        "when": "+1", "stayDays": 1, "lat": 54.0, "lon": 18.0,
        "notes": "nowy", "callType": "",
    })
    result = save(path, calls)
    saved = voyage(result)["calls"]
    assert [call["order"] for call in saved] == list(range(1, 7))
    assert saved[1]["name"] == "Nowy Port"
    assert saved[1]["callType"] == ""
    assert location_row(path, "Nowy Port")[2:5] == (54.0, 18.0, "Port")
    workbook = load_workbook(path, data_only=False)
    assert workbook["Etapy"]["A2"].value == "wynik pozostaje"


@pytest.mark.parametrize("call_type", ["Punkt_trasy", "Punkt_trasy_ukryty"])
def test_adds_each_route_point_type_with_zero_stay(call_type: str):
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    calls = copy_calls(path)
    calls.insert(2, {
        "sourceVisitId": "", "name": f"Nowy {call_type}", "country": "",
        "when": "+0", "stayDays": 0, "lat": 40.0, "lon": 4.0,
        "notes": "", "callType": call_type,
    })
    result = save(path, calls)
    added = voyage(result)["calls"][2]
    assert added["callType"] == call_type
    assert added["stayDays"] == 0
    assert location_row(path, added["name"])[2:4] == (40.0, 4.0)


def test_removes_call_without_deleting_location():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    calls = [call for call in copy_calls(path) if call["name"] != "Punkt A"]
    result = save(path, calls)
    assert "Punkt A" not in [call["name"] for call in voyage(result)["calls"]]
    assert location_row(path, "Punkt A")[5] == "zachowaj"


def test_reorders_and_assigns_contiguous_integer_order():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    calls = copy_calls(path)
    moved = calls.pop(2)
    calls.insert(1, moved)
    result = save(path, calls)
    saved = voyage(result)["calls"]
    assert [call["name"] for call in saved[:3]] == ["Triest", "Punkt A", "Barcelona"]
    assert [call["order"] for call in saved] == [1, 2, 3, 4, 5]


def test_editing_one_voyage_preserves_another_voyage():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    workbook = load_workbook(path)
    workbook["Rejsy"].append(["R2", "Drugi", "2024-02-01", "Czerwony", 4, None])
    porty = workbook["Porty"]
    porty.append(["R2", 1, "A", "Test", "0", 0, 1.0, 1.0, "drugi", None])
    porty.append([None, 2, "B", "Test", "+1", 1, 2.0, 2.0, "drugi", None])
    workbook.save(path)
    calls = copy_calls(path)
    calls.insert(1, {
        "sourceVisitId": "", "name": "Nowy", "country": "Test",
        "when": "+1", "stayDays": 1, "lat": 10.0, "lon": 10.0,
        "notes": "", "callType": "",
    })
    save(path, calls)
    document = load_port_document(path)
    second = next(item for item in document["voyages"] if item["id"] == "R2")
    assert [(call["name"], call["order"], call["notes"]) for call in second["calls"]] == [
        ("A", 1, "drugi"), ("B", 2, "drugi")
    ]


def test_rejects_two_coordinates_for_one_location_name():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    document = load_port_document(path)
    calls = [dict(item) for item in voyage(document)["calls"]]
    calls[1]["lat"], calls[1]["lon"] = 42.0, 3.0
    calls[3]["lat"], calls[3]["lon"] = 43.0, 4.0
    with pytest.raises(ValueError, match="jednej lokalizacji"):
        save_voyage_ports(path, "R1", calls, document["revision"])


def test_cancel_read_only_does_not_change_workbook():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    before = hashlib.sha256(path.read_bytes()).hexdigest()
    load_port_document(path)
    load_port_document(path)
    assert hashlib.sha256(path.read_bytes()).hexdigest() == before


def test_frontend_contains_port_tools_and_passage_draft_name_fix():
    port_app = (ROOT / "port-editor" / "app.js").read_text(encoding="utf-8")
    passage_app = (ROOT / "passage-editor" / "app.js").read_text(encoding="utf-8")
    assert "Punkt_trasy_ukryty" in port_app
    assert "Lokalizacje nie zostanie usunięte" in port_app
    assert "sameName.forEach" in port_app
    assert "map.panTo" in port_app
    assert "showDraftInSelect(current.name)" in passage_app
    assert 'status: "development"' in passage_app


def test_current_workbook_can_be_opened_and_saved_as_noop_on_copy():
    path = workdir() / "rejsy.xlsx"
    shutil.copy2(ROOT / "routes" / "rejsy.xlsx", path)
    document = load_port_document(path)
    current_voyage = voyage(document)
    before = path.read_bytes()
    result = save_voyage_ports(
        path, current_voyage["id"], current_voyage["calls"], document["revision"]
    )
    assert path.read_bytes() == before
    assert result["revision"] == document["revision"]
    assert not path.with_name("rejsy.ports-editor.bak.xlsx").exists()


def test_local_server_exposes_current_protocol_and_workbook():
    path = workdir() / "rejsy.xlsx"
    make_workbook(path)
    server = create_port_editor_server(path, ROOT, port=0)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        with urlopen(f"http://127.0.0.1:{server.server_port}/api/ports") as response:
            payload = json.loads(response.read().decode("utf-8"))
        assert payload["editorProtocol"] == EDITOR_PROTOCOL_VERSION
        assert payload["voyages"][0]["id"] == "R1"
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)
