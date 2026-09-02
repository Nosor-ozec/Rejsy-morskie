from __future__ import annotations

import hashlib
import json
import math
import re
import shutil
from collections import defaultdict
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from .excel_io import load_input
from .kml import NAMED_COLORS
from .models import PortCall

STATIC_FILES = ("index.html", "app.js", "style.css")
STATIC_DIRS = ("vendor",)
MEDIA_COLUMNS = {
    "Film_ID", "Dzien_od_portu", "Opis", "URL_Google_Drive", "Aktywny"
}
MEDIA_ID = re.compile(r"^(.+)_([0-9]+)$")


def build_local_site(
    rejsy_path: Path,
    media_path: Path,
    outputs_dir: Path,
    site_dir: Path,
    assets_dir: Path,
) -> list[Path]:
    """Buduje kompletną stronę lokalną z tych samych plików co publikacja."""

    if site_dir.resolve() == assets_dir.resolve():
        raise ValueError("Katalog podglądu musi być inny niż katalog źródeł strony")
    if site_dir.exists():
        shutil.rmtree(site_dir)
    site_dir.mkdir(parents=True)
    _copy_assets(assets_dir, site_dir)

    data_dir = site_dir / "data"
    if data_dir.exists():
        shutil.rmtree(data_dir)
    data_dir.mkdir(parents=True)

    route_data, bases_by_name = _route_data(rejsy_path, outputs_dir, data_dir)
    media_data = _media_data(media_path, rejsy_path, bases_by_name)
    _write_json(data_dir / "route.json", route_data)
    _write_json(data_dir / "media.json", media_data)

    manifest_path = site_dir / "build-manifest.json"
    manifest = {
        "schemaVersion": 1,
        "files": _hash_site_files(site_dir, exclude={manifest_path.name}),
    }
    _write_json(manifest_path, manifest)
    return [site_dir / name for name in STATIC_FILES] + [
        data_dir / "route.json", data_dir / "media.json", manifest_path
    ]


def publish_site(site_dir: Path, docs_dir: Path) -> list[Path]:
    """Kopiuje sprawdzony wynik lokalny do docs i weryfikuje jego sumy."""

    manifest_path = site_dir / "build-manifest.json"
    if not manifest_path.exists():
        raise ValueError(
            "Brak sprawdzonego podglądu. Najpierw uruchom Uruchom-Rejsy.cmd"
        )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    files = manifest.get("files")
    if not isinstance(files, dict) or not files:
        raise ValueError("Manifest podglądu jest niepoprawny")
    _verify_hashes(site_dir, files)

    docs_dir.mkdir(parents=True, exist_ok=True)
    for directory in ("data", "vendor"):
        target = docs_dir / directory
        if target.exists():
            shutil.rmtree(target)
    copied: list[Path] = []
    for relative_name in sorted(files):
        source = site_dir / Path(relative_name)
        target = docs_dir / Path(relative_name)
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)
        copied.append(target)
    shutil.copy2(manifest_path, docs_dir / manifest_path.name)
    _verify_hashes(docs_dir, files)
    return copied


def _copy_assets(assets_dir: Path, site_dir: Path) -> None:
    for name in STATIC_FILES:
        source = assets_dir / name
        if not source.exists():
            raise ValueError(f"Brak pliku strony: {source}")
        shutil.copy2(source, site_dir / name)
    for name in STATIC_DIRS:
        source = assets_dir / name
        if not source.is_dir():
            raise ValueError(f"Brak katalogu strony: {source}")
        target = site_dir / name
        if target.exists():
            shutil.rmtree(target)
        shutil.copytree(source, target)


def _route_data(
    rejsy_path: Path, outputs_dir: Path, data_dir: Path
) -> tuple[dict[str, object], dict[str, list[dict[str, object]]]]:
    voyages, calls = load_input(rejsy_path)
    ports = []
    route_points = []
    base_by_visit: dict[tuple[str, int], dict[str, object]] = {}
    bases_by_name: dict[str, list[dict[str, object]]] = defaultdict(list)
    for call in sorted(calls, key=lambda item: (item.voyage_id, item.order)):
        if call.lat is None or call.lon is None:
            raise ValueError(f"Port {call.port}: brak współrzędnych")
        key = _name_key(call.port)
        visit_key = (call.voyage_id, call.order)
        base = {"call": call, "leg": None, "baseFraction": 0.0}
        base_by_visit[visit_key] = base
        bases_by_name[key].append(base)
        item = {
            "voyageId": call.voyage_id,
            "order": call.order,
            "visitId": _visit_id(call),
            "name": call.port,
            "position": [call.lat, call.lon],
            "stayDays": call.stay_days,
        }
        if call.is_real_port:
            ports.append(item)
        elif call.is_visible_route_point:
            route_points.append(item)

    workbook = load_workbook(rejsy_path, data_only=True, read_only=True)
    sheet = workbook["Etapy"]
    headers = {
        str(cell.value).strip(): cell.column - 1 for cell in next(sheet.iter_rows())
        if cell.value is not None
    }
    required = {
        "Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu",
        "Dzien_od", "Dzien_do", "Zakres_dni", "Dystans_nm", "GeoJSON_path",
        "Status",
    }
    missing = required - headers.keys()
    if missing:
        raise ValueError(f"Arkusz Etapy: brak kolumn {sorted(missing)}")

    legs = []
    real_ports_by_voyage: dict[str, list[PortCall]] = defaultdict(list)
    calls_by_voyage: dict[str, list[PortCall]] = defaultdict(list)
    for call in sorted(calls, key=lambda item: (item.voyage_id, item.order)):
        calls_by_voyage[call.voyage_id].append(call)
        if call.is_real_port:
            real_ports_by_voyage[call.voyage_id].append(call)
    geojson_root = data_dir / "geojson"
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        if str(row[headers["Status"]] or "").casefold() != "gotowy":
            raise ValueError(
                f"Etap {row[headers['Nazwa_etapu']]} nie ma gotowej trasy GeoJSON"
            )
        relative_path = Path(str(row[headers["GeoJSON_path"]]))
        source = outputs_dir / relative_path
        feature = json.loads(source.read_text(encoding="utf-8"))
        geometry = feature.get("geometry")
        if not isinstance(geometry, dict) or geometry.get("type") != "LineString":
            raise ValueError(f"Niepoprawna geometria GeoJSON: {source}")
        lon_lat = geometry.get("coordinates")
        if not isinstance(lon_lat, list) or len(lon_lat) < 2:
            raise ValueError(f"Pusta geometria GeoJSON: {source}")
        coordinates = [[float(point[1]), float(point[0])] for point in lon_lat]
        public_geojson = geojson_root / relative_path
        public_geojson.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, public_geojson)
        day_from = int(row[headers["Dzien_od"]])
        day_to = int(row[headers["Dzien_do"]])
        voyage_id = str(row[headers["Rejs_ID"]])
        leg_number = int(row[headers["Etap_nr"]])
        real_ports = real_ports_by_voyage.get(voyage_id, [])
        if leg_number < 1 or leg_number >= len(real_ports):
            raise ValueError(
                f"Etap {leg_number} rejsu {voyage_id}: brak odpowiadających wizyt w Porty"
            )
        start_call = real_ports[leg_number - 1]
        end_call = real_ports[leg_number]
        if (
            str(row[headers["Port_start"]]) != start_call.port
            or str(row[headers["Port_koniec"]]) != end_call.port
        ):
            raise ValueError(
                f"Etap {leg_number} rejsu {voyage_id} nie odpowiada kolejności Porty"
            )
        leg = {
            "voyageId": voyage_id,
            "number": leg_number,
            "name": str(row[headers["Nazwa_etapu"]]),
            "startPort": str(row[headers["Port_start"]]),
            "endPort": str(row[headers["Port_koniec"]]),
            "startVisitId": _visit_id(start_call),
            "endVisitId": _visit_id(end_call),
            "days": str(row[headers["Zakres_dni"]]),
            "travelDays": day_to - day_from + 1,
            "distanceNm": row[headers["Dystans_nm"]],
            "geojson": (Path("data") / "geojson" / relative_path).as_posix(),
            "coordinates": coordinates,
        }
        legs.append(leg)
        start_base = base_by_visit[(start_call.voyage_id, start_call.order)]
        start_base["leg"] = leg
        start_base["baseFraction"] = 0.0
        for call in calls_by_voyage[voyage_id]:
            if (
                call.is_route_point
                and start_call.order < call.order < end_call.order
            ):
                fraction = _fraction_at_position(coordinates, [call.lat, call.lon])
                point_base = base_by_visit[(call.voyage_id, call.order)]
                point_base["leg"] = leg
                point_base["baseFraction"] = fraction

    voyage_rows = [
        {"id": item.voyage_id, "name": item.name, "color": _leaflet_color(item.route_color)}
        for item in voyages.values()
    ]
    return {
        "voyages": voyage_rows,
        "ports": ports,
        "routePoints": route_points,
        "legs": legs,
    }, dict(bases_by_name)


def _media_data(
    media_path: Path,
    rejsy_path: Path,
    bases_by_name: dict[str, list[dict[str, object]]],
) -> dict[str, object]:
    load_input(rejsy_path)
    workbook = load_workbook(media_path, data_only=True, read_only=True)
    if "Filmy" not in workbook.sheetnames:
        raise ValueError("Brak arkusza Filmy w media.xlsx")
    sheet = workbook["Filmy"]
    rows = sheet.iter_rows(values_only=True)
    headers = {str(value).strip(): index for index, value in enumerate(next(rows))}
    missing = MEDIA_COLUMNS - headers.keys()
    if missing:
        raise ValueError(f"Arkusz Filmy: brak kolumn {sorted(missing)}")

    result = []
    seen_ids: set[str] = set()
    for excel_row, row in enumerate(rows, start=2):
        if not any(value is not None for value in row):
            continue
        if str(row[headers["Aktywny"]] or "").strip().casefold() != "tak":
            continue
        media_id = str(row[headers["Film_ID"]] or "").strip()
        if not media_id or media_id.casefold() in seen_ids:
            raise ValueError(f"Filmy, wiersz {excel_row}: pusty lub powtórzony Film_ID")
        seen_ids.add(media_id.casefold())
        match = MEDIA_ID.fullmatch(media_id)
        if not match:
            raise ValueError(f"Filmy, wiersz {excel_row}: niepoprawny Film_ID {media_id}")
        media_base_name = match.group(1)
        bases = bases_by_name.get(_name_key(media_base_name), [])
        if not bases:
            raise ValueError(
                f"Filmy, wiersz {excel_row}: nieznany port lub punkt trasy w {media_id}"
            )
        description = str(row[headers["Opis"]] or "").strip()
        url = str(row[headers["URL_Google_Drive"]] or "").strip()
        parsed_url = urlparse(url)
        if not description or parsed_url.scheme not in {"http", "https"} or not parsed_url.netloc:
            raise ValueError(f"Filmy, wiersz {excel_row}: brak opisu lub niepoprawny URL")
        day = _day_value(row[headers["Dzien_od_portu"]], excel_row)
        # Media są własnością nazwy lokalizacji. Każda wizyta o tej nazwie
        # otrzymuje ten sam komplet mediów; historyczna Kolejnosc_wizyty jest
        # celowo ignorowana i nie wiąże danych z numeracją Porty.
        for base in bases:
            call = base["call"]
            if day == 0:
                position = [call.lat, call.lon]
                at_sea = call.call_type == "Punkt_trasy_ukryty"
            else:
                leg = base["leg"]
                if leg is None:
                    raise ValueError(
                        f"Filmy, wiersz {excel_row}: po {call.port} nie ma etapu"
                    )
                travel_days = float(leg["travelDays"])
                base_time = travel_days * float(base["baseFraction"])
                media_time = base_time + day
                if travel_days <= 0 or media_time > travel_days:
                    raise ValueError(
                        f"Filmy, wiersz {excel_row}: Dzien_od_portu przekracza czas etapu"
                    )
                position = _point_along(leg["coordinates"], media_time / travel_days)
                at_sea = True
            result.append({
                "id": media_id,
                "port": call.port,
                "base": call.port,
                "baseType": call.call_type or "Port",
                "baseVisitId": _visit_id(call),
                "visitOrder": call.order,
                "description": description,
                "url": url,
                "dayFromPort": day,
                "atSea": at_sea,
                "position": position,
            })
    return {"media": result}


def _name_key(value: str) -> str:
    return value.strip().casefold()


def _visit_id(call: PortCall) -> str:
    return f"{call.voyage_id}:{call.order}"


def _day_value(value: object, excel_row: int) -> float:
    if value in (None, ""):
        return 0.0
    try:
        result = float(str(value).strip().replace(",", "."))
    except ValueError as error:
        raise ValueError(
            f"Filmy, wiersz {excel_row}: Dzien_od_portu musi być liczbą"
        ) from error
    if result < 0:
        raise ValueError(f"Filmy, wiersz {excel_row}: Dzien_od_portu nie może być ujemny")
    return result


def _point_along(coordinates: list[list[float]], fraction: float) -> list[float]:
    fraction = max(0.0, min(1.0, fraction))
    lengths = [
        _haversine(coordinates[index - 1], coordinates[index])
        for index in range(1, len(coordinates))
    ]
    total = sum(lengths)
    if total <= 0:
        raise ValueError("GeoJSON etapu ma zerową długość")
    target = total * fraction
    travelled = 0.0
    for index, length in enumerate(lengths, start=1):
        if travelled + length >= target:
            part = 0.0 if length == 0 else (target - travelled) / length
            start, end = coordinates[index - 1], coordinates[index]
            return [
                start[0] + (end[0] - start[0]) * part,
                start[1] + (end[1] - start[1]) * part,
            ]
        travelled += length
    return list(coordinates[-1])


def _fraction_at_position(
    coordinates: list[list[float]], position: list[float]
) -> float:
    lengths = [
        _haversine(coordinates[index - 1], coordinates[index])
        for index in range(1, len(coordinates))
    ]
    total = sum(lengths)
    if total <= 0:
        raise ValueError("GeoJSON etapu ma zerową długość")
    closest_index = min(
        range(len(coordinates)),
        key=lambda index: _haversine(coordinates[index], position),
    )
    return sum(lengths[:closest_index]) / total


def _haversine(start: list[float], end: list[float]) -> float:
    lat1, lon1 = map(math.radians, start)
    lat2, lon2 = map(math.radians, end)
    dlat = lat2 - lat1
    dlon = (lon2 - lon1 + math.pi) % (2 * math.pi) - math.pi
    value = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * 6371.0088 * math.asin(math.sqrt(value))


def _leaflet_color(value: str) -> str:
    normalized = value.strip().casefold()
    color = NAMED_COLORS.get(normalized, value.removeprefix("#"))
    if len(color) != 6:
        raise ValueError("Kolor trasy musi być nazwą albo mieć format #RRGGBB")
    try:
        int(color, 16)
    except ValueError as error:
        raise ValueError("Kolor trasy musi być nazwą albo mieć format #RRGGBB") from error
    return f"#{color.upper()}"


def _write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def _hash_site_files(site_dir: Path, exclude: set[str]) -> dict[str, str]:
    result = {}
    for path in sorted(item for item in site_dir.rglob("*") if item.is_file()):
        relative = path.relative_to(site_dir).as_posix()
        if relative not in exclude:
            result[relative] = hashlib.sha256(path.read_bytes()).hexdigest()
    return result


def _verify_hashes(root: Path, files: dict[str, str]) -> None:
    for relative_name, expected in files.items():
        path = root / Path(relative_name)
        if not path.exists() or hashlib.sha256(path.read_bytes()).hexdigest() != expected:
            raise ValueError(f"Plik nie zgadza się ze sprawdzonym podglądem: {relative_name}")
