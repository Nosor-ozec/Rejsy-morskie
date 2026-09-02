from __future__ import annotations

import hashlib
import json
import math
import mimetypes
import os
import re
import shutil
import threading
import webbrowser
from copy import copy
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse
from uuid import uuid4

from openpyxl import load_workbook


PASSAGE_COLUMNS = {
    "Nazwa", "Kraj", "Lat", "Lon", "Typ", "Uwagi", "Przejscie",
    "Przejscie_lp", "Przejscie_status",
}
EDITOR_PROTOCOL_VERSION = 3
_WRITE_LOCK = threading.Lock()
_PASSAGE_NAME = re.compile(r"^[^\x00-\x1f\x7f]{1,100}$")


def _key(value: object) -> str:
    return str(value or "").strip().casefold()


def _revision(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _headers(sheet) -> dict[str, int]:
    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }
    missing = PASSAGE_COLUMNS - headers.keys()
    if missing:
        raise ValueError(
            "Arkusz Lokalizacje nie zawiera kolumn: " + ", ".join(sorted(missing))
        )
    return headers


def _coordinate(value: object, label: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{label} musi być liczbą")
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{label} musi być liczbą") from None
    if not math.isfinite(number):
        raise ValueError(f"{label} musi być liczbą skończoną")
    limit = 90 if label == "Lat" else 180
    if not -limit <= number <= limit:
        raise ValueError(f"{label} musi mieścić się w zakresie {-limit}..{limit}")
    return number


def _passage_name(value: object) -> str:
    name = str(value or "").strip()
    if not _PASSAGE_NAME.fullmatch(name):
        raise ValueError("Nazwa przejścia musi mieć od 1 do 100 czytelnych znaków")
    return name


def _passage_status(value: object) -> str:
    status = str(value or "").strip().casefold()
    if status not in {"development", "stable"}:
        raise ValueError("Przejscie_status musi mieć wartość development albo stable")
    return status


def _point_payload(raw: object, index: int, passage_name: str) -> dict[str, object]:
    if not isinstance(raw, dict):
        raise ValueError(f"Punkt {index} ma niepoprawną postać")
    location_type = raw.get("locationType")
    return {
        "name": f"{passage_name} {index:02d}",
        "order": index,
        "lat": _coordinate(raw.get("lat"), f"Punkt {index}: Lat"),
        "lon": _coordinate(raw.get("lon"), f"Punkt {index}: Lon"),
        "country": str(raw.get("country") or ""),
        "locationType": "Przejscie" if location_type is None else str(location_type),
        "notes": str(raw.get("notes") or ""),
    }


def _validate_location_names(sheet, headers: dict[str, int]) -> None:
    seen: dict[str, int] = {}
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, headers["Nazwa"]).value
        if value in (None, ""):
            continue
        normalized = _key(value)
        if normalized in seen:
            raise ValueError(
                f"Lokalizacje: Nazwa nie jest unikalna w wierszach "
                f"{seen[normalized]} i {row}: {value}"
            )
        seen[normalized] = row


def _read_document(path: Path) -> dict[str, object]:
    # Część poprawnych plików XLSX nie zawiera opcjonalnego wymiaru arkusza.
    # W trybie read-only openpyxl zwraca wtedy max_row=None. Zwykły odczyt
    # nadal nie zapisuje ani nie zmienia skoroszytu, a poprawnie wyznacza zakres.
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        if "Lokalizacje" not in workbook.sheetnames:
            raise ValueError("Brak arkusza Lokalizacje")
        sheet = workbook["Lokalizacje"]
        headers = _headers(sheet)
        groups: dict[str, dict[str, object]] = {}
        seen_location_names: dict[str, int] = {}
        for row in range(2, sheet.max_row + 1):
            location_name = sheet.cell(row, headers["Nazwa"]).value
            if location_name not in (None, ""):
                normalized_location = _key(location_name)
                if normalized_location in seen_location_names:
                    raise ValueError(
                        f"Lokalizacje: Nazwa nie jest unikalna w wierszach "
                        f"{seen_location_names[normalized_location]} i {row}: "
                        f"{location_name}"
                    )
                seen_location_names[normalized_location] = row
            passage_value = sheet.cell(row, headers["Przejscie"]).value
            order_value = sheet.cell(row, headers["Przejscie_lp"]).value
            status_value = sheet.cell(row, headers["Przejscie_status"]).value
            if (
                passage_value in (None, "") and order_value in (None, "")
                and status_value in (None, "")
            ):
                continue
            if passage_value in (None, "") and status_value not in (None, ""):
                raise ValueError(
                    f"Lokalizacje, wiersz {row}: Przejscie_status można podać "
                    "tylko dla punktu przejścia"
                )
            if passage_value in (None, "") or order_value in (None, ""):
                raise ValueError(
                    f"Lokalizacje, wiersz {row}: Przejscie i Przejscie_lp "
                    "muszą być podane razem"
                )
            passage_name = _passage_name(passage_value)
            status = _passage_status(status_value or "stable")
            try:
                order_number = float(order_value)
            except (TypeError, ValueError):
                order_number = -1.0
            if (
                isinstance(order_value, bool)
                or not order_number.is_integer()
                or order_number < 1
            ):
                raise ValueError(
                    f"Lokalizacje, wiersz {row}: Przejscie_lp musi być "
                    "dodatnią liczbą całkowitą"
                )
            group_key = _key(passage_name)
            group = groups.setdefault(
                group_key, {
                    "name": passage_name, "status": status, "rows": [],
                    "points": [],
                }
            )
            if group["name"] != passage_name:
                raise ValueError(
                    f"Niejednolita pisownia nazwy przejścia: "
                    f"{group['name']} / {passage_name}"
                )
            if group["status"] != status:
                raise ValueError(
                    f"Przejście {passage_name}: wszystkie wiersze muszą mieć "
                    "ten sam Przejscie_status"
                )
            group["rows"].append(row)
            group["points"].append({
                "name": str(location_name or "").strip(),
                "order": int(order_number),
                "lat": _coordinate(sheet.cell(row, headers["Lat"]).value, "Lat"),
                "lon": _coordinate(sheet.cell(row, headers["Lon"]).value, "Lon"),
                "country": str(sheet.cell(row, headers["Kraj"]).value or ""),
                "locationType": str(sheet.cell(row, headers["Typ"]).value or ""),
                "notes": str(sheet.cell(row, headers["Uwagi"]).value or ""),
            })
        passages = []
        for group in groups.values():
            points = sorted(group["points"], key=lambda item: item["order"])
            orders = [item["order"] for item in points]
            if orders != list(range(1, len(points) + 1)):
                raise ValueError(
                    f"Przejście {group['name']}: Przejscie_lp musi tworzyć "
                    f"ciąg 1..{len(points)}; jest {orders}"
                )
            if len(points) < 2:
                raise ValueError(
                    f"Przejście {group['name']} musi zawierać co najmniej 2 punkty"
                )
            passages.append({
                "name": group["name"], "status": group["status"],
                "points": points,
            })
        passages.sort(key=lambda item: item["name"].casefold())
        return {"revision": _revision(path), "passages": passages}
    finally:
        workbook.close()


def load_passage_document(path: Path) -> dict[str, object]:
    path = Path(path)
    if not path.is_file():
        raise ValueError(f"Nie znaleziono skoroszytu: {path}")
    return _read_document(path)


def _excel_number(value: object) -> float:
    """Return the numeric value produced by openpyxl's strict Excel serializer."""
    return float(f"{float(value):.16g}")


def _verification_differences(
    expected: list[dict[str, object]],
    actual: list[dict[str, object]],
) -> list[str]:
    differences: list[str] = []
    if len(expected) != len(actual):
        differences.append(
            f"liczba punktów: oczekiwano {len(expected)}, odczytano {len(actual)}"
        )
    expected_orders = [point.get("order") for point in expected]
    actual_orders = [point.get("order") for point in actual]
    if expected_orders != actual_orders:
        differences.append(
            f"Przejscie_lp: oczekiwano {expected_orders}, odczytano {actual_orders}"
        )
    labels = {
        "name": "Nazwa", "country": "Kraj", "locationType": "Typ",
        "notes": "Uwagi",
    }
    for index, (wanted, received) in enumerate(zip(expected, actual), 1):
        point_label = str(wanted.get("name") or f"punkt {index}")
        for field in ("name", "country", "locationType", "notes"):
            if wanted.get(field) != received.get(field):
                differences.append(
                    f"{point_label}: {labels[field]} — oczekiwano "
                    f"{wanted.get(field)!r} ({type(wanted.get(field)).__name__}), "
                    f"odczytano {received.get(field)!r} "
                    f"({type(received.get(field)).__name__})"
                )
        for field, label in (("lat", "Lat"), ("lon", "Lon")):
            wanted_value = wanted.get(field)
            received_value = received.get(field)
            serialized_wanted: object = "nieprawidłowa liczba"
            try:
                serialized_wanted = _excel_number(wanted_value)
                same = serialized_wanted == _excel_number(received_value)
            except (TypeError, ValueError):
                same = False
            if not same:
                differences.append(
                    f"{point_label}: {label} — oczekiwano {wanted_value!r} "
                    f"({type(wanted_value).__name__}; po zapisie Excela "
                    f"{serialized_wanted!r}), odczytano {received_value!r} "
                    f"({type(received_value).__name__})"
                )
    return differences


def _same_points(left: list[dict[str, object]], right: list[dict[str, object]]) -> bool:
    return not _verification_differences(left, right)


def save_passage(
    path: Path,
    passage_name: object,
    raw_points: object,
    expected_revision: object,
    original_name: object | None = None,
    passage_status: object | None = None,
) -> dict[str, object]:
    path = Path(path)
    name = _passage_name(passage_name)
    if not isinstance(raw_points, list) or len(raw_points) < 2:
        raise ValueError("Przejście musi zawierać co najmniej 2 punkty")
    if len(raw_points) > 500:
        raise ValueError("Przejście nie może zawierać więcej niż 500 punktów")
    points = [_point_payload(raw, index, name) for index, raw in enumerate(raw_points, 1)]
    revision = str(expected_revision or "").strip().casefold()
    if not re.fullmatch(r"[0-9a-f]{64}", revision):
        raise ValueError("Brak prawidłowej wersji kontrolnej skoroszytu")

    with _WRITE_LOCK:
        if _revision(path) != revision:
            raise RuntimeError(
                "Skoroszyt zmienił się od otwarcia edytora. Użyj Anuluj/odśwież "
                "i ponów edycję."
            )
        current = _read_document(path)
        current_by_key = {_key(item["name"]): item for item in current["passages"]}
        source_key = _key(original_name) if original_name not in (None, "") else ""
        target_key = _key(name)
        if source_key:
            if source_key not in current_by_key:
                raise ValueError(f"Nie istnieje edytowane przejście: {original_name}")
            if target_key != source_key and target_key in current_by_key:
                raise ValueError(f"Przejście o nazwie {name} już istnieje")
            existing = current_by_key[source_key]
            status = _passage_status(passage_status or existing["status"])
            if (
                target_key == source_key and existing["status"] == status
                and _same_points(existing["points"], points)
            ):
                return current
        else:
            if target_key in current_by_key:
                raise ValueError(
                    f"Przejście {name} już istnieje; wybierz je z listy"
                )
            status = _passage_status(passage_status or "development")

        workbook = load_workbook(path, data_only=False)
        temp_path = path.with_name(
            f".{path.stem}.editor-{uuid4().hex}.tmp{path.suffix}"
        )
        backup_path = path.with_name(f"{path.stem}.editor.bak{path.suffix}")
        try:
            sheet = workbook["Lokalizacje"]
            headers = _headers(sheet)
            matching_rows = [
                row for row in range(2, sheet.max_row + 1)
                if _key(sheet.cell(row, headers["Przejscie"]).value) == source_key
            ] if source_key else []
            if matching_rows:
                template_row = matching_rows[0]
            else:
                template_row = max(
                    2,
                    max(
                        (row for row in range(2, sheet.max_row + 1)
                         if any(sheet.cell(row, column).value is not None
                                for column in range(1, sheet.max_column + 1))),
                        default=1,
                    ),
                )
            template_styles = [
                copy(sheet.cell(template_row, column)._style)
                for column in range(1, sheet.max_column + 1)
            ]
            template_height = sheet.row_dimensions[template_row].height
            target_rows = matching_rows[:len(points)]
            if len(points) > len(matching_rows):
                insertion_row = matching_rows[-1] + 1 if matching_rows else sheet.max_row + 1
                extra = len(points) - len(matching_rows)
                sheet.insert_rows(insertion_row, extra)
                target_rows.extend(range(insertion_row, insertion_row + extra))
            elif len(points) < len(matching_rows):
                for row in reversed(matching_rows[len(points):]):
                    sheet.delete_rows(row, 1)
            for row, point in zip(target_rows, points):
                sheet.row_dimensions[row].height = template_height
                for column, style in enumerate(template_styles, 1):
                    sheet.cell(row, column)._style = copy(style)
                values = {
                    "Nazwa": point["name"], "Kraj": point["country"],
                    "Lat": point["lat"], "Lon": point["lon"],
                    "Typ": point["locationType"], "Uwagi": point["notes"],
                    "Przejscie": name, "Przejscie_lp": point["order"],
                    "Przejscie_status": status,
                }
                for column_name, value in values.items():
                    sheet.cell(row, headers[column_name]).value = value
                sheet.cell(row, headers["Lat"]).number_format = "0.000000"
                sheet.cell(row, headers["Lon"]).number_format = "0.000000"
                sheet.cell(row, headers["Przejscie_lp"]).number_format = "0"
            _validate_location_names(sheet, headers)
            shutil.copy2(path, backup_path)
            workbook.save(temp_path)
            verified = _read_document(temp_path)
            saved = next(
                (item for item in verified["passages"] if _key(item["name"]) == target_key),
                None,
            )
            if saved is None:
                available = [item["name"] for item in verified["passages"]]
                raise RuntimeError(
                    f"Kontrola zapisanego przejścia {name} nie powiodła się: "
                    f"nie znaleziono Przejscie={name!r}; odczytano {available!r}"
                )
            differences = []
            if saved["name"] != name:
                differences.append(
                    f"Przejscie: oczekiwano {name!r}, odczytano {saved['name']!r}"
                )
            if saved["status"] != status:
                differences.append(
                    f"Przejscie_status: oczekiwano {status!r}, "
                    f"odczytano {saved['status']!r}"
                )
            differences.extend(_verification_differences(points, saved["points"]))
            if differences:
                raise RuntimeError(
                    f"Kontrola zapisanego przejścia {name} nie powiodła się. "
                    f"Pierwsza różnica: {differences[0]}"
                )
            os.replace(temp_path, path)
            return _read_document(path)
        finally:
            workbook.close()
            temp_path.unlink(missing_ok=True)


class PassageEditorServer(ThreadingHTTPServer):
    daemon_threads = True
    allow_reuse_address = True

    def __init__(self, address, handler, workbook_path: Path, editor_dir: Path, vendor_dir: Path):
        super().__init__(address, handler)
        self.workbook_path = workbook_path
        self.editor_dir = editor_dir
        self.vendor_dir = vendor_dir


class PassageEditorHandler(BaseHTTPRequestHandler):
    server: PassageEditorServer

    def log_message(self, format: str, *args: Any) -> None:
        print(f"Edytor: {format % args}")

    def _json(self, status: int, value: object) -> None:
        data = json.dumps(value, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def _error(self, status: int, error: Exception) -> None:
        self._json(status, {"error": str(error)})

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/passages":
            try:
                self._json(HTTPStatus.OK, {
                    **load_passage_document(self.server.workbook_path),
                    "editorProtocol": EDITOR_PROTOCOL_VERSION,
                })
            except (OSError, RuntimeError, ValueError) as error:
                self._error(HTTPStatus.BAD_REQUEST, error)
            return
        relative = unquote(parsed.path).lstrip("/") or "index.html"
        if relative.startswith("vendor/"):
            root = self.server.vendor_dir
            relative = relative.removeprefix("vendor/")
        else:
            root = self.server.editor_dir
        try:
            path = (root / relative).resolve()
            path.relative_to(root.resolve())
            if not path.is_file():
                raise FileNotFoundError(relative)
            data = path.read_bytes()
            content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
            if content_type.startswith("text/") or content_type in {"application/javascript", "application/json"}:
                content_type += "; charset=utf-8"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(data)
        except (FileNotFoundError, ValueError):
            self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        if urlparse(self.path).path != "/api/passages":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length < 2 or length > 1_000_000:
                raise ValueError("Nieprawidłowy rozmiar danych")
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            if not isinstance(payload, dict):
                raise ValueError("Nieprawidłowe dane zapisu")
            result = save_passage(
                self.server.workbook_path,
                payload.get("name"), payload.get("points"),
                payload.get("revision"), payload.get("originalName"),
                payload.get("status"),
            )
            self._json(HTTPStatus.OK, {
                **result,
                "editorProtocol": EDITOR_PROTOCOL_VERSION,
                "message": "Dane zapisane. Uruchom Aktualizuj-Przejscia-SeaRouter.cmd, aby zastosować przejście.",
            })
        except RuntimeError as error:
            self._error(HTTPStatus.CONFLICT, error)
        except (json.JSONDecodeError, OSError, ValueError) as error:
            self._error(HTTPStatus.BAD_REQUEST, error)


def create_editor_server(
    workbook_path: Path,
    project_root: Path,
    host: str = "127.0.0.1",
    port: int = 8766,
) -> PassageEditorServer:
    if host not in {"127.0.0.1", "localhost", "::1"}:
        raise ValueError("Edytor może nasłuchiwać wyłącznie lokalnie")
    project_root = Path(project_root).resolve()
    editor_dir = project_root / "passage-editor"
    vendor_dir = project_root / "docs" / "vendor"
    for required in (editor_dir / "index.html", editor_dir / "app.js", editor_dir / "style.css", vendor_dir / "leaflet" / "leaflet.css"):
        if not required.is_file():
            raise ValueError(f"Brak pliku edytora: {required}")
    return PassageEditorServer(
        (host, port), PassageEditorHandler, Path(workbook_path).resolve(),
        editor_dir, vendor_dir,
    )


def serve_passage_editor(
    workbook_path: Path,
    project_root: Path,
    host: str = "127.0.0.1",
    port: int = 8766,
    open_browser: bool = True,
) -> None:
    server = create_editor_server(workbook_path, project_root, host, port)
    url = f"http://{host}:{server.server_port}/"
    print(f"Edytor przejść: {url}")
    print("Zamknięcie: Ctrl+C w tym oknie.")
    if open_browser:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("Edytor zatrzymany.")
    finally:
        server.server_close()
