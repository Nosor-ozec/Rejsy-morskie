from __future__ import annotations

import hashlib
import json
import mimetypes
import os
import re
import shutil
import threading
import webbrowser
from copy import copy
from datetime import date, datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse
from uuid import uuid4

from openpyxl import load_workbook

from .excel_io import LOKALIZACJE_COLUMNS, PORTY_COLUMNS, ROUTE_POINT_TYPES, load_input
from .schedule import calculate_schedule, parse_when


EDITOR_PROTOCOL_VERSION = 1
LOCATION_HEADERS = ["Nazwa", "Kraj", "Lat", "Lon", "Typ", "Uwagi"]
CALL_TYPES = {"", *ROUTE_POINT_TYPES}
_WRITE_LOCK = threading.Lock()


def _key(value: object) -> str:
    return str(value or "").strip().casefold()


def _revision(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _headers(sheet, required: set[str]) -> dict[str, int]:
    result = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }
    missing = required - result.keys()
    if missing:
        raise ValueError(f"Arkusz {sheet.title}: brak kolumn {sorted(missing)}")
    return result


def _json_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _excel_float(value: float) -> float:
    return float(f"{value:.16g}")


def _number(value: object, label: str, minimum: float, maximum: float) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{label} musi być liczbą")
    try:
        result = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label} musi być liczbą") from error
    if not minimum <= result <= maximum:
        raise ValueError(f"{label} musi mieścić się w zakresie {minimum}..{maximum}")
    return result


def _integer(value: object, label: str, minimum: int = 0) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label} musi być liczbą całkowitą")
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label} musi być liczbą całkowitą") from error
    if not number.is_integer() or number < minimum:
        raise ValueError(f"{label} musi być liczbą całkowitą >= {minimum}")
    return int(number)


def _location_rows(workbook) -> tuple[dict[str, dict[str, object]], dict[str, int]]:
    if "Lokalizacje" not in workbook.sheetnames:
        raise ValueError("Brak arkusza Lokalizacje")
    sheet = workbook["Lokalizacje"]
    headers = _headers(sheet, LOKALIZACJE_COLUMNS)
    locations: dict[str, dict[str, object]] = {}
    row_numbers: dict[str, int] = {}
    for row in range(2, sheet.max_row + 1):
        if not any(sheet.cell(row, column).value is not None for column in range(1, sheet.max_column + 1)):
            continue
        name = str(sheet.cell(row, headers["Nazwa"]).value or "").strip()
        if not name:
            raise ValueError(f"Lokalizacje, wiersz {row}: Nazwa nie może być pusta")
        key = _key(name)
        if key in locations:
            raise ValueError(f"Lokalizacje: powtórzona Nazwa po normalizacji: {name}")
        lat = _number(sheet.cell(row, headers["Lat"]).value, f"{name}.Lat", -90, 90)
        lon = _number(sheet.cell(row, headers["Lon"]).value, f"{name}.Lon", -180, 180)
        locations[key] = {
            "name": name,
            "country": str(sheet.cell(row, headers["Kraj"]).value or "").strip(),
            "lat": lat,
            "lon": lon,
            "locationType": str(sheet.cell(row, headers["Typ"]).value or "").strip(),
            "notes": str(sheet.cell(row, headers["Uwagi"]).value or "").strip(),
        }
        row_numbers[key] = row
    return locations, row_numbers


def _raw_calls(workbook) -> dict[tuple[str, int], dict[str, object]]:
    sheet = workbook["Porty"]
    headers = _headers(sheet, PORTY_COLUMNS)
    result: dict[tuple[str, int], dict[str, object]] = {}
    voyage_id: str | None = None
    for row in range(2, sheet.max_row + 1):
        entered = sheet.cell(row, headers["Rejs_ID"]).value
        if entered not in (None, ""):
            voyage_id = str(entered).strip()
        order = sheet.cell(row, headers["Kolejnosc"]).value
        if voyage_id is None or order in (None, ""):
            continue
        result[(voyage_id, int(order))] = {
            "when": _json_value(sheet.cell(row, headers["Kiedy"]).value),
            "storedLat": sheet.cell(row, headers["Lat"]).value,
            "storedLon": sheet.cell(row, headers["Lon"]).value,
        }
    return result


def load_port_document(path: Path) -> dict[str, object]:
    path = Path(path)
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        locations, _ = _location_rows(workbook)
        raw = _raw_calls(workbook)
        voyages, calls = load_input(path)
        voyage_items = []
        for voyage in voyages.values():
            voyage_calls = []
            for call in sorted(
                (item for item in calls if item.voyage_id == voyage.voyage_id),
                key=lambda item: item.order,
            ):
                source_id = f"{call.voyage_id}:{call.order}"
                raw_call = raw[(call.voyage_id, call.order)]
                voyage_calls.append({
                    "sourceVisitId": source_id,
                    "visitId": source_id,
                    "order": call.order,
                    "name": call.port,
                    "country": call.country or "",
                    "when": raw_call["when"],
                    "stayDays": call.stay_days,
                    "lat": call.lat,
                    "lon": call.lon,
                    "notes": call.notes or "",
                    "callType": call.call_type or "",
                    "coordinatesSource": call.coordinates_source or "",
                })
            voyage_items.append({
                "id": voyage.voyage_id,
                "name": voyage.name,
                "calls": voyage_calls,
            })
        return {
            "revision": _revision(path),
            "voyages": voyage_items,
            "locations": sorted(locations.values(), key=lambda item: _key(item["name"])),
        }
    finally:
        workbook.close()


def _call_payload(raw: object, index: int, voyage_id: str) -> dict[str, object]:
    if not isinstance(raw, dict):
        raise ValueError(f"Pozycja {index}: nieprawidłowe dane")
    name = str(raw.get("name") or "").strip()
    if not name:
        raise ValueError(f"Pozycja {index}: nazwa nie może być pusta")
    call_type = str(raw.get("callType") or "").strip()
    if call_type not in CALL_TYPES:
        raise ValueError(
            f"{name}: Typ musi być pusty, Punkt_trasy albo Punkt_trasy_ukryty"
        )
    when = raw.get("when")
    parse_when(when)
    stay_days = _integer(raw.get("stayDays", 1), f"{name}.Postoj_dni")
    if call_type and stay_days != 0:
        raise ValueError(f"{name}: punkt trasy musi mieć Postoj_dni=0")
    lat = _number(raw.get("lat"), f"{name}.Lat", -90, 90)
    lon = _number(raw.get("lon"), f"{name}.Lon", -180, 180)
    source_id = str(raw.get("sourceVisitId") or "").strip()
    if source_id and not re.fullmatch(re.escape(voyage_id) + r":\d+", source_id):
        raise ValueError(f"{name}: nieprawidłowy identyfikator źródłowej wizyty")
    return {
        "sourceVisitId": source_id,
        "order": index,
        "name": name,
        "country": str(raw.get("country") or "").strip(),
        "when": when,
        "stayDays": stay_days,
        "lat": lat,
        "lon": lon,
        "notes": str(raw.get("notes") or "").strip(),
        "callType": call_type,
    }


def _call_differences(expected: list[dict[str, object]], actual: list[dict[str, object]]) -> list[str]:
    if len(expected) != len(actual):
        return [f"liczba pozycji: oczekiwano {len(expected)}, odczytano {len(actual)}"]
    labels = {
        "order": "Kolejnosc", "name": "Port", "country": "Kraj",
        "when": "Kiedy", "stayDays": "Postoj_dni", "lat": "Lat",
        "lon": "Lon", "notes": "Uwagi", "callType": "Typ",
    }
    for index, (wanted, received) in enumerate(zip(expected, actual), 1):
        for field, label in labels.items():
            left, right = wanted[field], received[field]
            if field in {"lat", "lon"}:
                left = _excel_float(float(left))
                right = float(right)
            if left != right:
                return [
                    f"pozycja {index} ({wanted['name']}), {label}: "
                    f"oczekiwano {left!r}, odczytano {right!r}"
                ]
    return []


def _selected_voyage(document: dict[str, object], voyage_id: str) -> dict[str, object]:
    result = next((item for item in document["voyages"] if item["id"] == voyage_id), None)
    if result is None:
        raise ValueError(f"Nie istnieje rejs {voyage_id}")
    return result


def _same_call(existing: dict[str, object], edited: dict[str, object]) -> bool:
    fields = ("name", "country", "when", "stayDays", "lat", "lon", "notes", "callType")
    for field in fields:
        left, right = existing[field], edited[field]
        if field in {"lat", "lon"}:
            left, right = _excel_float(float(left)), _excel_float(float(right))
        if left != right:
            return False
    return True


def _voyage_rows(sheet, headers: dict[str, int], voyage_id: str) -> list[int]:
    result = []
    current: str | None = None
    for row in range(2, sheet.max_row + 1):
        entered = sheet.cell(row, headers["Rejs_ID"]).value
        if entered not in (None, ""):
            current = str(entered).strip()
        order = sheet.cell(row, headers["Kolejnosc"]).value
        if current == voyage_id and order not in (None, ""):
            result.append(row)
    return result


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, sheet.max_column + 1):
        sheet.cell(target_row, column)._style = copy(sheet.cell(source_row, column)._style)


def _resize_table(sheet, last_column: str) -> None:
    for table in sheet.tables.values():
        table.ref = f"A1:{last_column}{max(2, sheet.max_row)}"


def _upsert_locations(
    workbook,
    calls: list[dict[str, object]],
    names_to_write: set[str],
) -> None:
    sheet = workbook["Lokalizacje"]
    headers = _headers(sheet, LOKALIZACJE_COLUMNS)
    locations, rows_by_key = _location_rows(workbook)
    grouped: dict[str, list[dict[str, object]]] = {}
    for call in calls:
        if _key(call["name"]) in names_to_write:
            grouped.setdefault(_key(call["name"]), []).append(call)
    for key, same_name_calls in grouped.items():
        first = same_name_calls[0]
        expected = (_excel_float(float(first["lat"])), _excel_float(float(first["lon"])))
        for call in same_name_calls[1:]:
            received = (_excel_float(float(call["lat"])), _excel_float(float(call["lon"])))
            if received != expected:
                raise ValueError(
                    f"Wszystkie wizyty {first['name']} muszą używać jednej lokalizacji; "
                    f"odczytano różne Lat/Lon"
                )
        if key in rows_by_key:
            row = rows_by_key[key]
        else:
            row = sheet.max_row + 1
            template = max(2, sheet.max_row)
            _copy_row_style(sheet, template, row)
            sheet.cell(row, headers["Nazwa"]).value = first["name"]
            sheet.cell(row, headers["Kraj"]).value = first["country"] or None
            sheet.cell(row, headers["Typ"]).value = first["callType"] or "Port"
            sheet.cell(row, headers["Uwagi"]).value = None
        sheet.cell(row, headers["Lat"]).value = first["lat"]
        sheet.cell(row, headers["Lon"]).value = first["lon"]
        sheet.cell(row, headers["Lat"]).number_format = "0.000000"
        sheet.cell(row, headers["Lon"]).number_format = "0.000000"
    _resize_table(sheet, chr(64 + sheet.max_column))


def save_voyage_ports(
    path: Path,
    voyage_id: object,
    raw_calls: object,
    expected_revision: object,
) -> dict[str, object]:
    path = Path(path)
    selected_id = str(voyage_id or "").strip()
    if not selected_id:
        raise ValueError("Brak Rejs_ID")
    if not isinstance(raw_calls, list):
        raise ValueError("Nieprawidłowa lista pozycji")
    calls = [_call_payload(raw, index, selected_id) for index, raw in enumerate(raw_calls, 1)]
    if sum(not call["callType"] for call in calls) < 2:
        raise ValueError("Rejs musi zawierać co najmniej dwa zwykłe porty")
    revision = str(expected_revision or "").strip().casefold()
    if not re.fullmatch(r"[0-9a-f]{64}", revision):
        raise ValueError("Brak prawidłowej wersji kontrolnej skoroszytu")

    with _WRITE_LOCK:
        if _revision(path) != revision:
            raise RuntimeError(
                "Skoroszyt zmienił się od otwarcia edytora. Użyj Anuluj/odśwież i ponów edycję."
            )
        current = load_port_document(path)
        current_voyage = _selected_voyage(current, selected_id)
        existing_by_id = {
            call["sourceVisitId"]: call for call in current_voyage["calls"]
        }
        source_ids = [call["sourceVisitId"] for call in calls if call["sourceVisitId"]]
        if len(source_ids) != len(set(source_ids)):
            raise ValueError("Ta sama źródłowa wizyta występuje więcej niż raz")
        unknown = [source_id for source_id in source_ids if source_id not in existing_by_id]
        if unknown:
            raise ValueError(f"Nieznana źródłowa wizyta: {unknown[0]}")

        names_to_write: set[str] = set()
        for call in calls:
            source_id = call["sourceVisitId"]
            if not source_id:
                names_to_write.add(_key(call["name"]))
                continue
            existing = existing_by_id[source_id]
            if (
                _key(existing["name"]) != _key(call["name"])
                or _excel_float(float(existing["lat"])) != _excel_float(float(call["lat"]))
                or _excel_float(float(existing["lon"])) != _excel_float(float(call["lon"]))
            ):
                names_to_write.add(_key(call["name"]))

        normalized_existing = []
        for index, existing in enumerate(current_voyage["calls"], 1):
            item = {field: existing[field] for field in (
                "sourceVisitId", "order", "name", "country", "when", "stayDays",
                "lat", "lon", "notes", "callType",
            )}
            item["order"] = index
            normalized_existing.append(item)
        no_call_changes = len(calls) == len(normalized_existing) and all(
            edited["sourceVisitId"] == existing["sourceVisitId"]
            and _same_call(existing, edited)
            for edited, existing in zip(calls, normalized_existing)
        )
        if no_call_changes and not names_to_write:
            return current

        workbook = load_workbook(path, data_only=False)
        temp_path = path.with_name(f".{path.stem}.ports-editor-{uuid4().hex}.tmp{path.suffix}")
        backup_path = path.with_name(f"{path.stem}.ports-editor.bak{path.suffix}")
        try:
            sheet = workbook["Porty"]
            headers = _headers(sheet, PORTY_COLUMNS)
            matching_rows = _voyage_rows(sheet, headers, selected_id)
            if not matching_rows:
                raise ValueError(f"Porty: brak wierszy rejsu {selected_id}")
            template_row = matching_rows[0]
            target_rows = matching_rows[:len(calls)]
            if len(calls) > len(matching_rows):
                insertion = matching_rows[-1] + 1
                extra = len(calls) - len(matching_rows)
                sheet.insert_rows(insertion, extra)
                target_rows.extend(range(insertion, insertion + extra))
            elif len(calls) < len(matching_rows):
                for row in reversed(matching_rows[len(calls):]):
                    sheet.delete_rows(row, 1)
            for row, call in zip(target_rows, calls):
                _copy_row_style(sheet, template_row, row)
                values = {
                    "Rejs_ID": selected_id if call["order"] == 1 else None,
                    "Kolejnosc": call["order"], "Port": call["name"],
                    "Kraj": call["country"] or None, "Kiedy": call["when"],
                    "Postoj_dni": call["stayDays"], "Lat": call["lat"],
                    "Lon": call["lon"], "Uwagi": call["notes"] or None,
                    "Typ": call["callType"] or None,
                }
                for column_name, value in values.items():
                    sheet.cell(row, headers[column_name]).value = value
                sheet.cell(row, headers["Kiedy"]).number_format = "@"
                sheet.cell(row, headers["Lat"]).number_format = "0.000000"
                sheet.cell(row, headers["Lon"]).number_format = "0.000000"
            _resize_table(sheet, chr(64 + sheet.max_column))
            _upsert_locations(workbook, calls, names_to_write)
            shutil.copy2(path, backup_path)
            workbook.save(temp_path)

            verified = load_port_document(temp_path)
            verified_voyage = _selected_voyage(verified, selected_id)
            actual_calls = [{field: item[field] for field in (
                "order", "name", "country", "when", "stayDays", "lat", "lon",
                "notes", "callType",
            )} for item in verified_voyage["calls"]]
            differences = _call_differences(calls, actual_calls)
            if differences:
                raise RuntimeError(
                    f"Kontrola zapisu Porty nie powiodła się. Pierwsza różnica: {differences[0]}"
                )
            voyages, loaded_calls = load_input(temp_path)
            calculate_schedule(
                voyages[selected_id],
                [call for call in loaded_calls if call.voyage_id == selected_id],
            )
            os.replace(temp_path, path)
            return load_port_document(path)
        finally:
            workbook.close()
            temp_path.unlink(missing_ok=True)


class PortEditorServer(ThreadingHTTPServer):
    daemon_threads = True
    allow_reuse_address = True

    def __init__(self, address, handler, workbook_path: Path, editor_dir: Path, vendor_dir: Path):
        super().__init__(address, handler)
        self.workbook_path = workbook_path
        self.editor_dir = editor_dir
        self.vendor_dir = vendor_dir


class PortEditorHandler(BaseHTTPRequestHandler):
    server: PortEditorServer

    def log_message(self, format: str, *args: Any) -> None:
        print(f"Edytor portów: {format % args}")

    def _json(self, status: int, value: object) -> None:
        data = json.dumps(value, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/ports":
            try:
                self._json(HTTPStatus.OK, {
                    **load_port_document(self.server.workbook_path),
                    "editorProtocol": EDITOR_PROTOCOL_VERSION,
                })
            except (OSError, RuntimeError, ValueError) as error:
                self._json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return
        relative = unquote(parsed.path).lstrip("/") or "index.html"
        if relative.startswith("vendor/"):
            root = self.server.vendor_dir
            relative = relative.removeprefix("vendor/")
        else:
            root = self.server.editor_dir
        try:
            path = (root / relative).resolve()
            path.relative_to(root.resolve())
            if not path.is_file():
                raise FileNotFoundError(relative)
            data = path.read_bytes()
            content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
            if content_type.startswith("text/") or content_type == "application/javascript":
                content_type += "; charset=utf-8"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(data)
        except (FileNotFoundError, ValueError):
            self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        if urlparse(self.path).path != "/api/ports":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length < 2 or length > 2_000_000:
                raise ValueError("Nieprawidłowy rozmiar danych")
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            if not isinstance(payload, dict):
                raise ValueError("Nieprawidłowe dane zapisu")
            result = save_voyage_ports(
                self.server.workbook_path, payload.get("voyageId"),
                payload.get("calls"), payload.get("revision"),
            )
            self._json(HTTPStatus.OK, {
                **result,
                "editorProtocol": EDITOR_PROTOCOL_VERSION,
                "message": "Dane zapisane. Uruchom Uruchom-Rejsy.cmd, aby przeliczyć Etapy i mapę.",
            })
        except RuntimeError as error:
            self._json(HTTPStatus.CONFLICT, {"error": str(error)})
        except (json.JSONDecodeError, OSError, ValueError) as error:
            self._json(HTTPStatus.BAD_REQUEST, {"error": str(error)})


def create_port_editor_server(
    workbook_path: Path,
    project_root: Path,
    host: str = "127.0.0.1",
    port: int = 8767,
) -> PortEditorServer:
    if host not in {"127.0.0.1", "localhost", "::1"}:
        raise ValueError("Edytor może nasłuchiwać wyłącznie lokalnie")
    project_root = Path(project_root).resolve()
    editor_dir = project_root / "port-editor"
    vendor_dir = project_root / "docs" / "vendor"
    for required in (
        editor_dir / "index.html", editor_dir / "app.js", editor_dir / "style.css",
        vendor_dir / "leaflet" / "leaflet.css",
    ):
        if not required.is_file():
            raise ValueError(f"Brak pliku edytora: {required}")
    return PortEditorServer(
        (host, port), PortEditorHandler, Path(workbook_path).resolve(), editor_dir, vendor_dir
    )


def serve_port_editor(
    workbook_path: Path,
    project_root: Path,
    host: str = "127.0.0.1",
    port: int = 8767,
    open_browser: bool = True,
) -> None:
    server = create_port_editor_server(workbook_path, project_root, host, port)
    url = f"http://{host}:{server.server_port}/"
    print(f"Edytor portów i punktów rejsu: {url}")
    print("Zamknięcie: Ctrl+C w tym oknie.")
    if open_browser:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("Edytor zatrzymany.")
    finally:
        server.server_close()
