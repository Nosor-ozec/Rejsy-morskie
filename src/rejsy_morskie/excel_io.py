from __future__ import annotations

from pathlib import Path
from shutil import copy2
from typing import Iterable

from openpyxl import load_workbook

from .models import Leg, Location, PortCall, Voyage
from .schedule import parse_excel_date

REJSY_COLUMNS = {
    "Rejs_ID", "Nazwa_rejsu", "Data_startu", "Kolor_trasy", "CA", "Uwagi"
}
PORTY_COLUMNS = {
    "Rejs_ID", "Kolejnosc", "Port", "Kraj", "Kiedy", "Typ",
    "Postoj_dni", "Lat", "Lon", "Uwagi"
}
ROUTE_POINT_TYPES = {"Punkt_trasy", "Punkt_trasy_ukryty"}
LOKALIZACJE_COLUMNS = {"Nazwa", "Kraj", "Lat", "Lon", "Typ", "Uwagi"}
ETAPY_COLUMNS = [
    "Rejs_ID", "Etap_nr", "Port_start", "Port_koniec", "Nazwa_etapu",
    "Dzien_od", "Dzien_do", "Data_od", "Data_do", "Zakres_dni",
    "Zakres_dat", "Dystans_nm", "GeoJSON_path", "Status", "Uwagi",
]


def _rows(sheet) -> Iterable[tuple[int, dict[str, object]]]:
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(values)]
    except StopIteration:
        return
    for excel_row, row in enumerate(values, start=2):
        if any(value is not None for value in row):
            yield excel_row, dict(zip(headers, row))


def _require_columns(sheet, expected: set[str]) -> None:
    headers = {cell.value for cell in sheet[1]}
    missing = expected - headers
    if missing:
        raise ValueError(f"Arkusz {sheet.title}: brak kolumn {sorted(missing)}")


def _inherit_voyage_id(
    value: object, previous_voyage_id: str | None, excel_row: int
) -> str:
    entered_id = str(value).strip() if value is not None else ""
    if entered_id:
        return entered_id
    if previous_voyage_id is None:
        raise ValueError(
            f"Porty, wiersz {excel_row}: pierwszy Rejs_ID nie może być pusty"
        )
    return previous_voyage_id


def _normalize_location_name(value: object) -> str:
    return str(value).strip().casefold()


def _load_locations(workbook) -> dict[str, Location]:
    if "Lokalizacje" not in workbook.sheetnames:
        return {}
    sheet = workbook["Lokalizacje"]
    _require_columns(sheet, LOKALIZACJE_COLUMNS)
    locations: dict[str, Location] = {}
    rows_by_key: dict[str, int] = {}
    for excel_row, row in _rows(sheet):
        name = str(row["Nazwa"] or "").strip()
        if not name:
            raise ValueError(f"Lokalizacje, wiersz {excel_row}: Nazwa nie może być pusta")
        key = _normalize_location_name(name)
        if key in locations:
            raise ValueError(
                f"Lokalizacje, wiersze {rows_by_key[key]} i {excel_row}: "
                f"powtórzona Nazwa po normalizacji: {name}"
            )
        lat, lon = row["Lat"], row["Lon"]
        if lat is None or lon is None:
            raise ValueError(
                f"Lokalizacje, wiersz {excel_row} ({name}): Lat i Lon są obowiązkowe"
            )
        try:
            latitude, longitude = float(lat), float(lon)
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"Lokalizacje, wiersz {excel_row} ({name}): Lat i Lon muszą być liczbami"
            ) from error
        if not -90 <= latitude <= 90 or not -180 <= longitude <= 180:
            raise ValueError(
                f"Lokalizacje, wiersz {excel_row} ({name}): Lat/Lon poza zakresem"
            )
        locations[key] = Location(
            name=name,
            country=str(row["Kraj"]).strip() if row["Kraj"] is not None else None,
            lat=latitude,
            lon=longitude,
            location_type=(
                str(row["Typ"]).strip() if row["Typ"] is not None else None
            ),
            notes=str(row["Uwagi"]).strip() if row["Uwagi"] is not None else None,
        )
        rows_by_key[key] = excel_row
    return locations


def load_input(path: Path) -> tuple[dict[str, Voyage], list[PortCall]]:
    workbook = load_workbook(path)
    for name in ("Rejsy", "Porty", "Etapy"):
        if name not in workbook.sheetnames:
            raise ValueError(f"Brak arkusza {name}")

    _require_columns(workbook["Rejsy"], REJSY_COLUMNS)
    _require_columns(workbook["Porty"], PORTY_COLUMNS)
    locations = _load_locations(workbook)

    voyages: dict[str, Voyage] = {}
    for _, row in _rows(workbook["Rejsy"]):
        voyage_id = str(row["Rejs_ID"]).strip()
        if voyage_id in voyages:
            raise ValueError(f"Powtórzony Rejs_ID: {voyage_id}")
        voyages[voyage_id] = Voyage(
            voyage_id=voyage_id,
            name=str(row["Nazwa_rejsu"]).strip(),
            start_date=parse_excel_date(row["Data_startu"]),
            route_color=str(row["Kolor_trasy"] or "#0057B8").strip(),
            ca=str(row["CA"]).strip() if row["CA"] is not None else None,
            notes=str(row["Uwagi"]).strip() if row["Uwagi"] is not None else None,
        )

    calls: list[PortCall] = []
    inherited_voyage_id: str | None = None
    for excel_row, row in _rows(workbook["Porty"]):
        inherited_voyage_id = _inherit_voyage_id(
            row["Rejs_ID"], inherited_voyage_id, excel_row
        )
        voyage_id = inherited_voyage_id
        if voyage_id not in voyages:
            raise ValueError(
                f"Porty, wiersz {excel_row}: nieznany Rejs_ID {voyage_id}"
            )
        port_name = str(row["Port"]).strip()
        lat, lon = row["Lat"], row["Lon"]
        call_type = str(row["Typ"] or "").strip() or None
        if call_type is not None and call_type not in ROUTE_POINT_TYPES:
            raise ValueError(
                f"Porty, wiersz {excel_row}: Typ musi być pusty, "
                "Punkt_trasy albo Punkt_trasy_ukryty"
            )
        if (lat is None) != (lon is None):
            raise ValueError(f"Port {row['Port']}: Lat i Lon muszą występować razem")
        coordinates_source = "porty" if lat is not None else None
        location = locations.get(_normalize_location_name(port_name))
        if lat is None and location is not None:
            lat, lon = location.lat, location.lon
            coordinates_source = "lokalizacje"
        if call_type in ROUTE_POINT_TYPES and (lat is None or lon is None):
            raise ValueError(
                f"Punkt trasy {row['Port']}: Lat i Lon są obowiązkowe; "
                "brak Lat/Lon w Porty i wpisu w Lokalizacje. "
                "Punkty trasy nie są geokodowane"
            )
        stay_days = int(row["Postoj_dni"]) if row["Postoj_dni"] is not None else 1
        if call_type in ROUTE_POINT_TYPES and stay_days != 0:
            raise ValueError(
                f"Punkt trasy {row['Port']}: Postoj_dni musi wynosić 0"
            )
        calls.append(
            PortCall(
                voyage_id=voyage_id,
                order=int(row["Kolejnosc"]),
                port=port_name,
                country=str(row["Kraj"]).strip() if row["Kraj"] is not None else None,
                when=row["Kiedy"],
                stay_days=stay_days,
                lat=float(lat) if lat is not None else None,
                lon=float(lon) if lon is not None else None,
                notes=str(row["Uwagi"]).strip() if row["Uwagi"] is not None else None,
                call_type=call_type,
                coordinates_source=coordinates_source,
            )
        )
    return voyages, calls

def _save_workbook_atomic(workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    temp_path = output_path.with_name(
        f"{output_path.stem}.tmp{output_path.suffix}"
    )
    backup_path = output_path.with_name(
        f"{output_path.stem}.bak{output_path.suffix}"
    )

    if output_path.exists():
        copy2(output_path, backup_path)

    workbook.save(temp_path)
    temp_path.replace(output_path)


def write_results(
    input_path: Path,
    output_path: Path,
    calls: list[PortCall],
    legs: list[Leg],
) -> None:
    workbook = load_workbook(input_path)
    _write_port_coordinates(workbook["Porty"], calls)

    sheet = workbook["Etapy"]
    values = [ETAPY_COLUMNS]
    values.extend(
        [
            leg.voyage_id,
            leg.number,
            leg.start_port,
            leg.end_port,
            leg.name,
            leg.day_from,
            leg.day_to,
            leg.date_from,
            leg.date_to,
            leg.day_range,
            f"{leg.date_from.isoformat()} – {leg.date_to.isoformat()}",
            leg.distance_nm,
            str(leg.geojson_path) if leg.geojson_path else None,
            leg.status,
            leg.notes,
        ]
        for leg in legs
    )

    _replace_values_preserving_template(sheet, values)

    if sheet.tables:
        last_row = max(2, len(values))
        for table in sheet.tables.values():
            table.ref = f"A1:O{last_row}"

    _save_workbook_atomic(workbook, output_path)


def write_legs(input_path: Path, output_path: Path, legs: list[Leg]) -> None:
    """Zachowany interfejs dla wcześniejszego polecenia `schedule`."""

    workbook = load_workbook(input_path)
    sheet = workbook["Etapy"]
    values = [ETAPY_COLUMNS]
    values.extend(
        [
            leg.voyage_id, leg.number, leg.start_port, leg.end_port, leg.name,
            leg.day_from, leg.day_to, leg.date_from, leg.date_to, leg.day_range,
            f"{leg.date_from.isoformat()} – {leg.date_to.isoformat()}",
            leg.distance_nm, str(leg.geojson_path) if leg.geojson_path else None,
            leg.status, leg.notes,
        ]
        for leg in legs
    )
    _replace_values_preserving_template(sheet, values)
    if sheet.tables:
        last_row = max(2, len(values))
        for table in sheet.tables.values():
            table.ref = f"A1:O{last_row}"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _replace_values_preserving_template(sheet, values: list[list[object]]) -> None:
    """Podmienia dane, ale zostawia szerokości, style i formaty szablonu."""

    required_rows = len(values)
    if required_rows > sheet.max_row:
        template_row = max(2, sheet.max_row)
        for row_number in range(sheet.max_row + 1, required_rows + 1):
            for column in range(1, len(values[0]) + 1):
                source = sheet.cell(template_row, column)
                target = sheet.cell(row_number, column)
                target._style = source._style
                target.number_format = source.number_format

    for row in sheet.iter_rows(min_col=1, max_col=len(values[0])):
        for cell in row:
            cell.value = None
    for row_number, row_values in enumerate(values, start=1):
        for column, value in enumerate(row_values, start=1):
            sheet.cell(row_number, column).value = value


def _write_port_coordinates(sheet, calls: list[PortCall]) -> None:
    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}
    call_by_key = {(call.voyage_id, call.order): call for call in calls}
    previous_voyage_id: str | None = None
    for row_number in range(2, sheet.max_row + 1):
        entered_id = sheet.cell(row_number, headers["Rejs_ID"]).value
        if entered_id not in (None, ""):
            previous_voyage_id = str(entered_id).strip()
        order = sheet.cell(row_number, headers["Kolejnosc"]).value
        if previous_voyage_id is None or order in (None, ""):
            continue
        call = call_by_key.get((previous_voyage_id, int(order)))
        if call is None:
            continue
        if call.coordinates_source == "porty":
            continue
        if call.lat is not None and call.lon is not None:
            sheet.cell(row_number, headers["Lat"]).value = call.lat
            sheet.cell(row_number, headers["Lon"]).value = call.lon
