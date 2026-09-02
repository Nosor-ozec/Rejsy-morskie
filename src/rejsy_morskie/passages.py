from __future__ import annotations

import json
import math
import os
import re
import tempfile
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class PassagePoint:
    name: str
    order: int
    latitude: float
    longitude: float


def _slug(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", ascii_value.casefold()).strip("-")


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object, label: str, row: int) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"Lokalizacje, wiersz {row}: {label} musi być liczbą") from None
    if not math.isfinite(number):
        raise ValueError(f"Lokalizacje, wiersz {row}: {label} musi być liczbą skończoną")
    return number


def read_passages(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        if "Lokalizacje" not in workbook.sheetnames:
            raise ValueError("Brak arkusza Lokalizacje")
        sheet = workbook["Lokalizacje"]
        headers = {_text(cell.value): index for index, cell in enumerate(sheet[1], 1)}
        required = {"Nazwa", "Lat", "Lon", "Przejscie", "Przejscie_lp"}
        missing = sorted(required - headers.keys())
        if missing:
            raise ValueError(
                "Arkusz Lokalizacje nie zawiera kolumn: " + ", ".join(missing)
            )

        status_column = headers.get("Przejscie_status")
        groups: dict[str, dict[str, object]] = {}
        for row in range(2, sheet.max_row + 1):
            passage = _text(sheet.cell(row, headers["Przejscie"]).value)
            order_value = sheet.cell(row, headers["Przejscie_lp"]).value
            order_text = _text(order_value)
            status_text = _text(sheet.cell(row, status_column).value) if status_column else ""
            if not passage and not order_text and not status_text:
                continue
            if not passage and status_text:
                raise ValueError(
                    f"Lokalizacje, wiersz {row}: Przejscie_status można podać tylko dla punktu przejścia"
                )
            if not passage or not order_text:
                raise ValueError(
                    f"Lokalizacje, wiersz {row}: Przejscie i Przejscie_lp muszą być podane razem"
                )
            if isinstance(order_value, bool):
                raise ValueError(f"Lokalizacje, wiersz {row}: Przejscie_lp musi być dodatnią liczbą całkowitą")
            try:
                order_float = float(order_value)
            except (TypeError, ValueError):
                raise ValueError(f"Lokalizacje, wiersz {row}: Przejscie_lp musi być dodatnią liczbą całkowitą") from None
            if not order_float.is_integer() or order_float < 1:
                raise ValueError(f"Lokalizacje, wiersz {row}: Przejscie_lp musi być dodatnią liczbą całkowitą")
            latitude = _number(sheet.cell(row, headers["Lat"]).value, "Lat", row)
            longitude = _number(sheet.cell(row, headers["Lon"]).value, "Lon", row)
            if not -90 <= latitude <= 90 or not -180 <= longitude <= 180:
                raise ValueError(f"Lokalizacje, wiersz {row}: współrzędne są poza dozwolonym zakresem")
            name = _text(sheet.cell(row, headers["Nazwa"]).value)
            if not name:
                raise ValueError(f"Lokalizacje, wiersz {row}: Nazwa jest wymagana")
            status = status_text.casefold() or "stable"
            if status not in {"development", "stable"}:
                raise ValueError(
                    f"Lokalizacje, wiersz {row}: Przejscie_status musi mieć wartość development albo stable"
                )
            key = passage.casefold()
            group = groups.setdefault(
                key, {"name": passage, "status": status, "points": []}
            )
            display_name = str(group["name"])
            if passage != display_name and passage.casefold() == display_name.casefold():
                passage = display_name
            if group["status"] != status:
                raise ValueError(
                    f"Przejście {display_name}: wszystkie wiersze muszą mieć ten sam Przejscie_status"
                )
            group["points"].append(
                PassagePoint(name, int(order_float), latitude, longitude)
            )

        result: list[dict] = []
        for key in sorted(groups):
            group = groups[key]
            name = str(group["name"])
            status = str(group["status"])
            points = group["points"]
            points.sort(key=lambda point: point.order)
            orders = [point.order for point in points]
            if len(set(orders)) != len(orders):
                raise ValueError(f"Przejście {name}: Przejscie_lp musi być unikalne")
            expected = list(range(1, len(points) + 1))
            if orders != expected:
                raise ValueError(
                    f"Przejście {name}: Przejscie_lp musi tworzyć ciąg 1..{len(points)}; jest {orders}"
                )
            if len(points) < 2:
                raise ValueError(f"Przejście {name}: wymagane są co najmniej dwa punkty")
            node_count = len(points)
            result.append(
                {
                    "id": _slug(name),
                    "name": name,
                    "status": status,
                    "origin": "routes/rejsy.xlsx:Lokalizacje",
                    "reason": "Ręcznie zatwierdzone punkty przejścia używane wyłącznie do routingu i wizualizacji, nie do nawigacji.",
                    "waypoints": [[point.longitude, point.latitude] for point in points],
                    "injection": {
                        "node_depth": 1.0,
                        "chain_consecutive_waypoints": True,
                        "endpoint_nearest_nodes": 5,
                        "maximum_endpoint_connection_km": 100.0,
                    },
                    "observed_graph_delta": {
                        "nodes": node_count,
                        "edges": node_count - 1 + 10,
                    },
                }
            )
        return result
    finally:
        workbook.close()


def generate_passages(workbook_path: Path, output_path: Path) -> list[dict]:
    passages = read_passages(workbook_path)
    document = {
        "schema_version": 1,
        "coordinate_order": "longitude_latitude",
        "generated_from": "routes/rejsy.xlsx:Lokalizacje",
        "passages": passages,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(
        prefix=output_path.name + ".", suffix=".tmp", dir=output_path.parent
    )
    try:
        with os.fdopen(handle, "w", encoding="utf-8", newline="\n") as stream:
            json.dump(document, stream, ensure_ascii=False, indent=2)
            stream.write("\n")
        os.replace(temporary_name, output_path)
    except BaseException:
        Path(temporary_name).unlink(missing_ok=True)
        raise
    return passages
